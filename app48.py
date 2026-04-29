import streamlit as st
import pandas as pd
import math
import sqlite3
import json
import os
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()
GROQ_API_KEY      = os.getenv('GROQ_API_KEY', '')
ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY', '')

# Si estamos en Streamlit Cloud, leer desde secrets
try:
    import streamlit as _st
    _secrets = _st.secrets._secrets if hasattr(_st.secrets, '_secrets') else {}
    if 'GROQ_API_KEY' in _secrets and not GROQ_API_KEY:
        GROQ_API_KEY = _secrets['GROQ_API_KEY']
    if 'ANTHROPIC_API_KEY' in _secrets and not ANTHROPIC_API_KEY:
        ANTHROPIC_API_KEY = _secrets['ANTHROPIC_API_KEY']
except Exception:
    pass

# ─────────────────────────────────────────────
# FIREBASE
# ─────────────────────────────────────────────
FIREBASE_KEY = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'aprov-c526a-firebase-adminsdk-fbsvc-c7b86e52ba.json')

def get_firestore():
    import firebase_admin
    from firebase_admin import credentials, firestore
    if not firebase_admin._apps:
        try:
            # Streamlit Cloud: leer desde secrets
            _has_firebase_secret = False
            try:
                _has_firebase_secret = 'firebase' in st.secrets
            except Exception:
                pass
            if _has_firebase_secret:
                import json
                cert = dict(st.secrets['firebase'])
                # Streamlit convierte \n en literales, hay que restaurarlos
                if 'private_key' in cert:
                    cert['private_key'] = cert['private_key'].replace('\\n', '\n')
                cred = credentials.Certificate(cert)
            # Local: leer desde archivo .json
            elif os.path.exists(FIREBASE_KEY):
                cred = credentials.Certificate(FIREBASE_KEY)
            else:
                return None, "No se encontró configuración de Firebase"
            firebase_admin.initialize_app(cred)
        except Exception as e:
            return None, str(e)
    try:
        db = firestore.client()
        return db, None
    except Exception as e:
        return None, str(e)

def firebase_guardar(coleccion, doc_id, datos):
    """Guarda un diccionario en Firestore."""
    db, err = get_firestore()
    if err:
        return False, err
    try:
        db.collection(coleccion).document(doc_id).set(datos)
        return True, None
    except Exception as e:
        return False, str(e)

def firebase_leer(coleccion, doc_id):
    """Lee un documento de Firestore. Devuelve dict o None."""
    db, err = get_firestore()
    if err:
        return None, err
    try:
        doc = db.collection(coleccion).document(doc_id).get()
        if doc.exists:
            return doc.to_dict(), None
        return None, "No existe"
    except Exception as e:
        return None, str(e)

def firebase_borrar(coleccion, doc_id):
    """Borra un documento de Firestore."""
    db, err = get_firestore()
    if err:
        return False, err
    try:
        db.collection(coleccion).document(doc_id).delete()
        return True, None
    except Exception as e:
        return False, str(e)

def df_a_firebase(df, coleccion, doc_id):
    """Guarda un DataFrame en Firestore dividido en chunks si es necesario."""
    try:
        db, err = get_firestore()
        if err:
            return False, err
        json_str = df.to_json(orient='records', date_format='iso', default_handler=str)
        chunk_size = 900000
        chunks = [json_str[i:i+chunk_size] for i in range(0, len(json_str), chunk_size)]
        # Borrar chunks anteriores
        meta_old = db.collection(coleccion).document(doc_id).get()
        if meta_old.exists:
            n_old = meta_old.to_dict().get('chunks', 1)
            for i in range(n_old):
                db.collection(coleccion).document(f'{doc_id}_chunk_{i}').delete()
        # Guardar metadata y chunks
        db.collection(coleccion).document(doc_id).set({'doc_id': doc_id, 'chunks': int(len(chunks))})
        for i, chunk in enumerate(chunks):
            db.collection(coleccion).document(f'{doc_id}_chunk_{i}').set({'doc_id': doc_id, 'chunk': i, 'data': chunk})
        return True, None
    except Exception as e:
        return False, str(e)

def firebase_a_df(coleccion, doc_id):
    """Lee un DataFrame de Firestore ensamblando chunks."""
    try:
        db, err = get_firestore()
        if err:
            return None, err

        meta = db.collection(coleccion).document(doc_id).get()
        if not meta.exists:
            return None, "No existe"

        meta_dict = meta.to_dict()

        # Formato antiguo: datos directamente en campo 'data'
        if 'data' in meta_dict and 'chunks' not in meta_dict:
            df = pd.read_json(meta_dict['data'], orient='records')
            return df, None

        # Formato nuevo: chunks separados
        n_chunks = int(str(meta_dict.get('chunks', 1)))
        json_parts = []
        for i in range(n_chunks):
            chunk_doc = db.collection(coleccion).document(f'{doc_id}_chunk_{i}').get()
            if chunk_doc.exists:
                json_parts.append(chunk_doc.to_dict().get('data', ''))

        if not json_parts:
            return None, "Sin chunks"

        import io
        df = pd.read_json(io.StringIO(''.join(json_parts)), orient='records')
        return df, None
    except Exception as e:
        return None, str(e)

def firebase_borrar_df(coleccion, doc_id):
    """Borra un DataFrame y todos sus chunks de Firestore."""
    try:
        db, err = get_firestore()
        if err:
            return False, err
        meta = db.collection(coleccion).document(doc_id).get()
        if meta.exists:
            n_chunks = meta.to_dict().get('chunks', 1)
            for i in range(n_chunks):
                db.collection(coleccion).document(f'{doc_id}_chunk_{i}').delete()
            db.collection(coleccion).document(doc_id).delete()
        return True, None
    except Exception as e:
        return False, str(e)

st.set_page_config(page_title="Aprovisionamiento Aldelis", layout="wide")

# ─────────────────────────────────────────────
# CONTROL DE ACCESO
# ─────────────────────────────────────────────
def check_password():
    if 'autenticado' not in st.session_state:
        st.session_state.autenticado = False
    if not st.session_state.autenticado:
        st.image("https://www.aldelis.es/wp-content/uploads/2021/03/logo-aldelis.png", width=200) if False else None
        st.title("🔒 Aprovisionamiento Aldelis")
        pwd = st.text_input("Contraseña:", type="password")
        if st.button("Entrar"):
            password_correcta = os.getenv("APP_PASSWORD", "aldelis2025")
            try:
                if hasattr(st, 'secrets') and 'APP_PASSWORD' in st.secrets._secrets:
                    password_correcta = st.secrets['APP_PASSWORD']
            except Exception:
                pass
            if pwd == password_correcta:
                st.session_state.autenticado = True
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")
        st.stop()

check_password()

# ─────────────────────────────────────────────
# COLUMNAS ESPERADAS (fuente de verdad única)
# ─────────────────────────────────────────────
COL_MAESTRO   = ['Referencia', 'Descripcion', 'Lead_time', 'Stock_seguridad', 'Unidades_palet', 'Incremento']
COL_STOCK     = ['Referencia', 'Almacen', 'Cantidad']
COL_CONSUMOS  = ['Referencia', 'Fecha', 'Cantidad']

ALMACENES_INT   = {'AL6', 'AL6SGA', 'AL6 SGA'}
ALMACENES_MERCA = {'ARENTO', 'CAMARA BANDEJAS F19'}
ALMACENES_TXT   = {'TXT'}

# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
if 'df_final' not in st.session_state:
    st.session_state.df_final = None
if 'df_transito' not in st.session_state:
    st.session_state.df_transito = pd.DataFrame(columns=['Referencia', 'Cantidad'])
if 'df_consumos' not in st.session_state:
    st.session_state.df_consumos = None
if 'firebase_cargado' not in st.session_state:
    st.session_state.firebase_cargado = False

# ── Inicializar session state ANTES de cargar Firebase ───────
if 'df_final' not in st.session_state:
    st.session_state.df_final = None
if 'df_transito' not in st.session_state:
    st.session_state.df_transito = pd.DataFrame(columns=['Referencia', 'Cantidad'])
if 'df_consumos' not in st.session_state:
    st.session_state.df_consumos = None

# ── Carga automática desde Firebase al arrancar ──────────────
if not st.session_state.firebase_cargado:
    _fb_errores = []
    try:
        df_fb, err = firebase_a_df('bandejas', 'df_final')
        if df_fb is not None:
            st.session_state.df_final = df_fb
        elif err and err not in ("No existe",) and not err.startswith('[{'):
            _fb_errores.append(f"bandejas/df_final: {err}")
        df_cons_fb, err = firebase_a_df('bandejas', 'df_consumos')
        if df_cons_fb is not None:
            df_cons_fb['Fecha'] = pd.to_datetime(df_cons_fb['Fecha'], errors='coerce')
            st.session_state.df_consumos = df_cons_fb
        df_mat_fb, _ = firebase_a_df('materiales', 'df_materiales')
        if df_mat_fb is not None:
            st.session_state.df_materiales = df_mat_fb
        df_etq_fb, _ = firebase_a_df('etiquetas', 'df_etiquetas_final')
        if df_etq_fb is not None:
            st.session_state.df_etiquetas_final = df_etq_fb
        df_vent_fb, _ = firebase_a_df('etiquetas', 'df_ventas')
        if df_vent_fb is not None:
            st.session_state.df_ventas = df_vent_fb

        # Tránsitos
        for key, coleccion in [
            ('df_transito',     'bandejas'),
            ('df_transito2',    'bandejas'),
            ('df_transito_etq', 'etiquetas'),
        ]:
            df_t_fb, _ = firebase_a_df(coleccion, key)
            if df_t_fb is not None:
                st.session_state[key] = df_t_fb

        # Pedidos
        df_ped_fb, _ = firebase_a_df('pedidos', 'df_pedidos')
        if df_ped_fb is not None:
            df_ped_fb['Fecha_entrega'] = pd.to_datetime(df_ped_fb['Fecha_entrega'], errors='coerce')
            st.session_state.df_pedidos = df_ped_fb

        # Planificación
        df_plan_fb, _ = firebase_a_df('planificacion', 'df_planificacion')
        if df_plan_fb is not None:
            st.session_state.df_planificacion = df_plan_fb

        # Paletización
        df_pal_fb, _ = firebase_a_df('logistica', 'df_paletizacion')
        if df_pal_fb is not None:
            st.session_state.df_paletizacion = df_pal_fb

        # Producto Terminado
        df_spt_fb, _ = firebase_a_df('producto_terminado', 'df_stock_pt')
        if df_spt_fb is not None:
            st.session_state.df_stock_pt = df_spt_fb
        df_ppt_fb, _ = firebase_a_df('producto_terminado', 'df_produccion_pt')
        if df_ppt_fb is not None:
            st.session_state.df_produccion_pt = df_ppt_fb
        df_pp_fb, _ = firebase_a_df('producto_terminado', 'df_plan_produccion')
        if df_pp_fb is not None:
            st.session_state.df_plan_produccion = df_pp_fb

        # Envase (etiquetas de caja)
        df_env_fb, _ = firebase_a_df('etiquetas', 'df_envase')
        if df_env_fb is not None:
            st.session_state.df_envase = df_env_fb
    except Exception as e:
        _fb_errores.append(str(e))
    if _fb_errores:
        st.sidebar.warning(f"⚠️ Firebase: {'; '.join(_fb_errores)}")
    elif st.session_state.df_final is not None:
        st.sidebar.success("☁️ Datos cargados desde Firebase")
    st.session_state.firebase_cargado = True

if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []
if 'df_materiales' not in st.session_state:
    st.session_state.df_materiales = None
if 'df_etiquetas_final' not in st.session_state:
    st.session_state.df_etiquetas_final = None
if 'df_ventas' not in st.session_state:
    st.session_state.df_ventas = None
if 'df_transito_etq' not in st.session_state:
    st.session_state.df_transito_etq = pd.DataFrame(columns=['Referencia', 'Cantidad'])
if 'df_transito2' not in st.session_state:
    st.session_state.df_transito2 = pd.DataFrame(columns=['Referencia', 'Cantidad'])
if 'df_pedidos' not in st.session_state:
    st.session_state.df_pedidos = pd.DataFrame(columns=['Referencia', 'Cantidad', 'Fecha_entrega'])
if 'df_planificacion' not in st.session_state:
    st.session_state.df_planificacion = None
if 'df_stock_pt' not in st.session_state:
    st.session_state.df_stock_pt = None
if 'df_produccion_pt' not in st.session_state:
    st.session_state.df_produccion_pt = None
if 'df_plan_produccion' not in st.session_state:
    st.session_state.df_plan_produccion = None
if 'logistica_historial' not in st.session_state:
    st.session_state.logistica_historial = []
if 'logistica_archivos' not in st.session_state:
    st.session_state.logistica_archivos = {}
if 'df_paletizacion' not in st.session_state:
    st.session_state.df_paletizacion = None
if 'df_envase' not in st.session_state:
    st.session_state.df_envase = None

# ─────────────────────────────────────────────
# BASE DE DATOS SQLITE
# ─────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'datos.db')

def init_db():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS snapshots (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha     TEXT,
            referencia TEXT,
            descripcion TEXT,
            stock_interno REAL,
            stock_merca   REAL,
            stock_txt     REAL,
            en_transito   REAL,
            cdm           REAL,
            stock_seguridad REAL,
            unidades_palet  REAL
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS consumos_diarios (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha     TEXT,
            referencia TEXT,
            cantidad   REAL
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS stock_pt_historico (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha      TEXT,
            referencia TEXT,
            descripcion TEXT,
            cantidad   REAL
        )
    ''')
    con.commit()
    con.close()

def guardar_snapshot(df_final, df_consumos):
    init_db()
    fecha = datetime.now().strftime('%Y-%m-%d')
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    # Borrar snapshot del mismo día si ya existe
    cur.execute("DELETE FROM snapshots WHERE fecha = ?", (fecha,))
    for _, row in df_final.iterrows():
        cur.execute('''INSERT INTO snapshots
            (fecha, referencia, descripcion, stock_interno, stock_merca, stock_txt,
             en_transito, cdm, stock_seguridad, unidades_palet)
            VALUES (?,?,?,?,?,?,?,?,?,?)''', (
            fecha,
            row.get('Referencia',''), row.get('Descripcion',''),
            row.get('Stock_interno',0), row.get('Stock_merca',0), row.get('Stock_txt',0),
            0,  # tránsito se añade en tiempo real
            row.get('Cdm',0), row.get('Stock_seguridad',0), row.get('Unidades_palet',1)
        ))
    # Guardar consumos diarios
    if df_consumos is not None:
        cur.execute("DELETE FROM consumos_diarios WHERE fecha >= (SELECT MIN(fecha) FROM consumos_diarios)")
        for _, row in df_consumos.iterrows():
            cur.execute('''INSERT INTO consumos_diarios (fecha, referencia, cantidad)
                VALUES (?,?,?)''', (
                str(row['Fecha'].date()) if hasattr(row['Fecha'], 'date') else str(row['Fecha']),
                row['Referencia'], float(row['Cantidad'])
            ))
    con.commit()
    con.close()

def obtener_contexto_agente():
    """Construye un contexto compacto para el agente (respeta límite de tokens de Groq free)."""
    init_db()
    con = sqlite3.connect(DB_PATH)
    lineas = []

    # Últimas fechas
    fechas = pd.read_sql("SELECT DISTINCT fecha FROM snapshots ORDER BY fecha DESC LIMIT 2", con)
    fechas_list = fechas['fecha'].tolist()

    if fechas_list:
        ultima = fechas_list[0]
        penultima = fechas_list[1] if len(fechas_list) > 1 else None
        lineas.append(f"FECHA ACTUAL: {ultima}")

        # Stock hoy — solo campos clave
        hoy = pd.read_sql(f"SELECT referencia, descripcion, stock_interno, stock_merca, stock_txt, cdm, stock_seguridad, unidades_palet FROM snapshots WHERE fecha='{ultima}'", con)
        hoy['u_p'] = hoy['unidades_palet'].clip(lower=1)
        hoy['pal_int']   = (hoy['stock_interno'] / hoy['u_p']).round()
        hoy['pal_merca'] = (hoy['stock_merca']   / hoy['u_p']).round()
        hoy['pal_txt']   = (hoy['stock_txt']      / hoy['u_p']).round()
        hoy['cdm_pal']   = hoy['cdm'].apply(lambda x: round(x,1))
        hoy['seg']       = hoy['stock_seguridad'].round()

        lineas.append("STOCK POR REFERENCIA (palets): ref | desc | int | merca | txt | cdm/dia | seg")
        for _, r in hoy.iterrows():
            lineas.append(f"  {r['referencia']} | {str(r['descripcion'])[:30]} | {int(r['pal_int'])} | {int(r['pal_merca'])} | {int(r['pal_txt'])} | {r['cdm_pal']} | {int(r['seg'])}")

        # Comparativa ayer
        if penultima:
            ayer = pd.read_sql(f"SELECT referencia, stock_interno, unidades_palet FROM snapshots WHERE fecha='{penultima}'", con)
            ayer['u_p'] = ayer['unidades_palet'].clip(lower=1)
            ayer['pal_int_ayer'] = (ayer['stock_interno'] / ayer['u_p']).round()
            comp = hoy[['referencia','pal_int']].merge(ayer[['referencia','pal_int_ayer']], on='referencia', how='left')
            comp['diff'] = comp['pal_int'] - comp['pal_int_ayer'].fillna(0)
            bajadas = comp[comp['diff'] < -2].sort_values('diff')
            if not bajadas.empty:
                lineas.append(f"BAJADAS DE STOCK RESPECTO A {penultima} (>2 pal):")
                for _, r in bajadas.iterrows():
                    lineas.append(f"  {r['referencia']}: {int(r['diff'])} pal")

    # Picos por día de semana (top referencias)
    consumos = pd.read_sql("SELECT referencia, fecha, cantidad FROM consumos_diarios", con)
    if not consumos.empty:
        consumos['fecha'] = pd.to_datetime(consumos['fecha'])
        consumos['dia'] = consumos['fecha'].dt.day_name()
        consumos['cantidad'] = consumos['cantidad'].abs()
        picos = consumos.groupby(['referencia','dia'])['cantidad'].mean().reset_index()
        # Solo top 15 referencias por consumo total
        top_refs = consumos.groupby('referencia')['cantidad'].sum().nlargest(15).index
        picos = picos[picos['referencia'].isin(top_refs)]
        lineas.append("PICOS POR DÍA DE SEMANA (media unidades, top 15 refs):")
        for ref, grp in picos.groupby('referencia'):
            dias_str = ', '.join([f"{r['dia'][:3]}:{int(r['cantidad'])}" for _, r in grp.iterrows()])
            lineas.append(f"  {ref}: {dias_str}")

    con.close()

    # Tránsito
    if not st.session_state.df_transito.empty:
        lineas.append("EN TRÁNSITO:")
        for _, r in st.session_state.df_transito.iterrows():
            lineas.append(f"  {r['Referencia']}: {r['Cantidad']} ud")

    # Producto Terminado
    if st.session_state.df_stock_pt is not None:
        lineas.append("\nPRODUCTO TERMINADO:")
        df_pt_ctx = st.session_state.df_stock_pt.copy()
        df_pt_ctx['Referencia'] = df_pt_ctx['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
        df_pt_ctx['Cantidad'] = pd.to_numeric(df_pt_ctx['Cantidad'], errors='coerce').fillna(0)
        if st.session_state.df_ventas is not None:
            v_ctx = st.session_state.df_ventas.copy()
            v_ctx['Referencia'] = v_ctx['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
            v_ctx['Unidades'] = pd.to_numeric(v_ctx['Unidades'], errors='coerce').fillna(0)
            vmd_ctx = (v_ctx.groupby('Referencia')['Unidades'].sum() / 22).reset_index()
            vmd_ctx.columns = ['Referencia', 'VMD']
            df_pt_ctx = df_pt_ctx.merge(vmd_ctx, on='Referencia', how='left')
            df_pt_ctx['VMD'] = df_pt_ctx['VMD'].fillna(0)
            df_pt_ctx['Dias'] = df_pt_ctx.apply(lambda r: round(r['Cantidad']/r['VMD'],1) if r['VMD']>0 else 999, axis=1)
            peligro_ctx = df_pt_ctx[df_pt_ctx['Dias'] < 1]
            lineas.append(f"  En peligro (<1 dia): {len(peligro_ctx)} referencias")
            for _, r in peligro_ctx.head(10).iterrows():
                lineas.append(f"  {r['Referencia']} stock:{int(r['Cantidad'])} VMD:{round(r['VMD'])} dias:{r['Dias']}")

    # Sugerencia turno noche
    try:
        con_pt = sqlite3.connect(DB_PATH)
        hist_pt = pd.read_sql("SELECT referencia, AVG(cantidad) as stock_medio FROM stock_pt_historico GROUP BY referencia", con_pt)
        con_pt.close()
        if not hist_pt.empty and st.session_state.df_ventas is not None:
            hist_pt['referencia'] = hist_pt['referencia'].str.upper().str.strip()
            vmd_n = vmd_ctx.rename(columns={'Referencia': 'referencia'}) if 'vmd_ctx' in dir() else pd.DataFrame()
            if not vmd_n.empty:
                hist_pt = hist_pt.merge(vmd_n, on='referencia', how='left')
                hist_pt['VMD'] = hist_pt['VMD'].fillna(0)
                hist_pt['ratio'] = hist_pt.apply(lambda r: round(r['VMD']/max(r['stock_medio'],1),4), axis=1)
                top_noche = hist_pt[hist_pt['VMD']>0].sort_values('ratio', ascending=False).head(15)
                lineas.append("\nSUGERENCIA TURNO NOCHE:")
                for _, r in top_noche.iterrows():
                    lineas.append(f"  {r['referencia']} stock_medio:{round(r['stock_medio'])} VMD:{round(r['VMD'])} ratio:{r['ratio']}")
    except Exception:
        pass

    return '\n'.join(lineas)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def leer_excel(file, nombre):
    """
    Lee un Excel detectando automáticamente la fila de cabeceras,
    por si hay filas vacías o de título encima de los datos.
    """
    try:
        raw = pd.read_excel(file, header=None)
        # Buscar primera fila que tenga algún texto (esa es la cabecera)
        header_row = 0
        for i, row in raw.iterrows():
            valores = row.dropna().tolist()
            if len(valores) > 0 and any(isinstance(v, str) for v in valores):
                header_row = i
                break
        df = pd.read_excel(file, header=header_row)
        return df
    except Exception as e:
        st.error(f"Error al leer '{nombre}': {e}")
        return None


def normalizar_columnas(df):
    """
    Quita espacios y renombra columnas al nombre canónico (case-insensitive).
    Así 'Stock_Seguridad', 'stock_seguridad' o 'STOCK_SEGURIDAD' funcionan igual.
    """
    df.columns = [str(c).strip() for c in df.columns]

    canonicos = {
        'referencia':      'Referencia',
        'descripcion':     'Descripcion',
        'lead_time':       'Lead_time',
        'stock_seguridad': 'Stock_seguridad',
        'unidades_palet':  'Unidades_palet',
        'incremento':      'Incremento',
        'almacen':         'Almacen',
        'cantidad':        'Cantidad',
        'fecha':           'Fecha',
    }
    rename_map = {
        col: canonicos[col.lower()]
        for col in df.columns
        if col.lower() in canonicos and col != canonicos[col.lower()]
    }
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def columnas_faltantes(df, requeridas, nombre_archivo):
    """Comprueba columnas. Devuelve lista de las que faltan."""
    faltantes = [c for c in requeridas if c not in df.columns]
    if faltantes:
        st.error(
            f"❌ **{nombre_archivo}** le faltan columnas: `{'`, `'.join(faltantes)}`\n\n"
            f"Columnas detectadas: `{'`, `'.join(df.columns.tolist())}`"
        )
    return faltantes


# ─────────────────────────────────────────────
# NAVEGACIÓN
# ─────────────────────────────────────────────
st.sidebar.title("🎮 Panel de Control")
st.sidebar.caption("📦 v2.0 | app.py")
menu = st.sidebar.radio("Selecciona:", ["📂 Cargar Archivos", "📊 Dashboard", "📈 Análisis", "🤖 Agente IA", "🔗 Materiales", "🏷️ Etiquetas", "🚢 Tránsito", "📋 Pedidos", "🔍 Previsión y Obsoletos", "🏪 Producto Terminado", "🏭 Planificación Producción", "🧠 Logística AI"])

# ══════════════════════════════════════════════
# MÓDULO 1: CARGAR ARCHIVOS
# ══════════════════════════════════════════════
if menu == "📂 Cargar Archivos":
    st.header("📂 Importación de Datos")

    st.info(
        "**Columnas requeridas:**\n"
        "- **Maestro**: `Referencia`, `Descripcion`, `Lead_time`, `Stock_seguridad`, `Unidades_palet`\n"
        "- **Stock**: `Referencia`, `Almacen`, `Cantidad`\n"
        "- **Consumos**: `Referencia`, `Fecha`, `Cantidad`"
    )

    f_maestro  = st.file_uploader("1. Maestro Artículos (.xlsx)", type="xlsx")
    f_stock    = st.file_uploader("2. Stock Actual (.xlsx)", type="xlsx")
    f_consumos = st.file_uploader("3. Histórico Consumos (.xlsx)", type="xlsx")

    # Archivo de paletización permanente
    with st.expander("📦 Archivo de Paletización (envases)"):
        f_pal_main = st.file_uploader("Subir/actualizar paletización (.xlsx)", type="xlsx", key="fpal_main")
        if f_pal_main and st.button("💾 Guardar paletización", key="btn_pal_main"):
            df_pal_main = pd.read_excel(f_pal_main)
            df_pal_main = normalizar_columnas(df_pal_main)
            st.session_state.df_paletizacion = df_pal_main
            df_a_firebase(df_pal_main, 'logistica', 'df_paletizacion')
            st.success(f"✅ {len(df_pal_main)} envases guardados en Firebase.")
            st.rerun()
        if st.session_state.df_paletizacion is not None:
            st.info(f"✅ {len(st.session_state.df_paletizacion)} envases cargados.")
            st.dataframe(st.session_state.df_paletizacion.iloc[:,[0]], use_container_width=True)
        else:
            st.warning("Sin archivo de paletización. Súbelo aquí para que persista.")

    if st.button("🚀 Sincronizar"):
        if not (f_maestro and f_stock and f_consumos):
            st.error("Sube los 3 archivos antes de sincronizar.")
            st.stop()

        # --- Leer ---
        m = leer_excel(f_maestro, "Maestro")
        s = leer_excel(f_stock, "Stock")
        c = leer_excel(f_consumos, "Consumos")
        if m is None or s is None or c is None:
            st.stop()

        # --- Normalizar cabeceras ---
        m = normalizar_columnas(m)
        s = normalizar_columnas(s)
        c = normalizar_columnas(c)

        # --- Validar columnas ---
        errores = (
            columnas_faltantes(m, COL_MAESTRO,  "Maestro") +
            columnas_faltantes(s, COL_STOCK,    "Stock") +
            columnas_faltantes(c, COL_CONSUMOS, "Consumos")
        )
        if errores:
            st.stop()

        # --- Limpiar Referencia (string sin espacios, mayúsculas) ---
        for df in [m, s, c]:
            df['Referencia'] = df['Referencia'].astype(str).str.strip().str.upper()

        # --- Procesar Stock por almacén ---
        s['Almacen'] = s['Almacen'].astype(str).str.strip()
        s['Cantidad'] = pd.to_numeric(s['Cantidad'], errors='coerce').fillna(0)

        res_stock = (
            s.groupby('Referencia')
             .apply(lambda g: pd.Series({
                 'Stock_interno': g.loc[g['Almacen'].isin(ALMACENES_INT),   'Cantidad'].sum(),
                 'Stock_merca':   g.loc[g['Almacen'].isin(ALMACENES_MERCA), 'Cantidad'].sum(),
                 'Stock_txt':     g.loc[g['Almacen'].isin(ALMACENES_TXT),   'Cantidad'].sum(),
             }))
             .reset_index()
        )

        # --- Procesar Consumos: CDM = media de palets consumidos en días con movimiento ---
        c['Fecha'] = pd.to_datetime(c['Fecha'], errors='coerce')
        c['Cantidad'] = pd.to_numeric(c['Cantidad'], errors='coerce').fillna(0).abs()
        m['Unidades_palet'] = pd.to_numeric(m['Unidades_palet'], errors='coerce').fillna(1)

        # Truncar fecha a día (ignorar hora) antes de agrupar
        c['Fecha'] = c['Fecha'].dt.normalize()

        # Consumo por referencia y día (en unidades), solo días con movimiento
        c_dia = c.groupby(['Referencia', 'Fecha'])['Cantidad'].sum().reset_index()
        c_dia = c_dia[c_dia['Cantidad'] > 0]

        # Unir Unidades_palet del maestro para convertir a palets por día
        # Normalizamos Referencia a mayúsculas en ambos lados para evitar fallos de merge
        uds_palet = m[['Referencia', 'Unidades_palet']].drop_duplicates().copy()
        uds_palet['Referencia'] = uds_palet['Referencia'].str.upper().str.strip()
        c_dia['Referencia'] = c_dia['Referencia'].str.upper().str.strip()
        c_dia = c_dia.merge(uds_palet, on='Referencia', how='left')
        c_dia['Unidades_palet'] = c_dia['Unidades_palet'].fillna(1).clip(lower=1)
        c_dia['Palets_dia'] = c_dia['Cantidad'] / c_dia['Unidades_palet']

        # Media de palets por día (solo días con movimiento)
        cdm = (
            c_dia.groupby('Referencia')['Palets_dia']
             .mean()
             .reset_index()
             .rename(columns={'Palets_dia': 'Cdm'})
        )

        # --- DEBUG TEMPORAL: mostrar C12043 ---
        ref_debug = 'C12043'
        st.write("### 🔍 DEBUG C12043")
        c12_raw = c[c['Referencia'] == ref_debug]
        st.write(f"Filas en consumos: {len(c12_raw)}")
        st.write(f"Cantidad total (abs): {c12_raw['Cantidad'].sum()}")
        c12_dia = c_dia[c_dia['Referencia'] == ref_debug]
        st.write(f"Días con movimiento: {len(c12_dia)}")
        st.write(c12_dia[['Fecha','Cantidad','Unidades_palet','Palets_dia']])
        c12_cdm = cdm[cdm['Referencia'] == ref_debug]
        st.write(f"CDM resultado: {c12_cdm['Cdm'].values}")

        # --- Quedarse solo con las columnas necesarias del Maestro ---
        m = m[COL_MAESTRO]

        # --- Unión: solo referencias presentes en Maestro Y Stock ---
        final = pd.merge(m, res_stock, on='Referencia', how='inner')
        final = pd.merge(final, cdm,      on='Referencia', how='left')

        # Rellenar numéricos con 0 donde no haya consumos o datos
        for col in ['Cdm', 'Stock_interno', 'Stock_merca', 'Stock_txt', 'Lead_time', 'Stock_seguridad', 'Unidades_palet', 'Incremento']:
            if col in final.columns:
                final[col] = pd.to_numeric(final[col], errors='coerce').fillna(0)

        st.session_state.df_final = final
        st.session_state.df_consumos = c
        guardar_snapshot(final, c)

        # Guardar en Firebase
        with st.spinner("Guardando en Firebase..."):
            ok1, e1 = df_a_firebase(final, 'bandejas', 'df_final')
            ok2, e2 = df_a_firebase(c,     'bandejas', 'df_consumos')
            if ok1 and ok2:
                st.success(f"✅ ¡Sincronizado! {len(final)} referencias guardadas y respaldadas en Firebase.")
            else:
                st.warning(f"⚠️ Sincronizado localmente pero error en Firebase: {e1 or e2}")
        st.dataframe(final.head(10), use_container_width=True)


# ══════════════════════════════════════════════
# MÓDULO 2: DASHBOARD
# ══════════════════════════════════════════════
elif menu == "📊 Dashboard":
    st.header("📊 Aprovisionamiento Aldelis")

    if st.session_state.df_final is None:
        st.warning("⚠️ Primero carga y sincroniza los archivos.")
        st.stop()

    df = st.session_state.df_final.copy()

    # Añadir tránsito
    t = (
        st.session_state.df_transito
        .groupby('Referencia')['Cantidad']
        .sum()
        .reset_index()
        .rename(columns={'Cantidad': 'En_transito'})
    )
    df = df.merge(t, on='Referencia', how='left')
    df['En_transito'] = df['En_transito'].fillna(0)

    # Tránsito 2
    t2 = (
        st.session_state.df_transito2
        .groupby('Referencia')['Cantidad'].sum()
        .reset_index()
        .rename(columns={'Cantidad': 'En_transito2'})
    )
    df = df.merge(t2, on='Referencia', how='left')
    df['En_transito2'] = df['En_transito2'].fillna(0)
    if 'Stock_avitrans' not in df.columns:
        df['Stock_avitrans'] = 0

    # --- Calcular palets y alerta ---
    def calcular_alerta(row):
        u_p        = max(row['Unidades_palet'], 1)
        lead       = row['Lead_time']
        seg        = row['Stock_seguridad']   # ya en palets
        cdm        = row['Cdm']               # ya en palets/día
        incremento = row.get('Incremento', 0) or 0

        pal_int       = round(row['Stock_interno']         / u_p)
        pal_merca     = round(row['Stock_merca']           / u_p)
        pal_txt       = round(row['Stock_txt']             / u_p)
        pal_avitrans  = round(row.get('Stock_avitrans', 0) / u_p)
        pal_transito  = round(row['En_transito']           / u_p)
        pal_transito2 = round(row.get('En_transito2', 0)   / u_p)
        seg_pal       = round(seg)
        cdm_pal       = math.ceil(cdm)

        # Stock disponible cuando llegue el nuevo pedido:
        # interno + tránsitos (llegan hoy) - consumo durante lead_time
        necesidad_bruta = cdm * lead
        stock_final = pal_int + pal_transito + pal_transito2 - necesidad_bruta

        # Alerta: no llegamos al stock de seguridad
        if stock_final < seg:
            pedido = math.ceil(seg - stock_final + incremento)
            dias_stock_actual = (pal_int / cdm) if cdm > 0 else 999
            if dias_stock_actual < lead:
                return pal_int, pal_merca, pal_txt, pal_avitrans, pal_transito, pal_transito2, seg_pal, cdm_pal, f"🔴 COMPRAR: {pedido} Pal.", "#721c24"
            else:
                return pal_int, pal_merca, pal_txt, pal_avitrans, pal_transito, pal_transito2, seg_pal, cdm_pal, f"🟡 COMPRAR: {pedido} Pal.", "#856404"

        # Sin necesidad de pedir -> siempre verde
        msg = "🟢 OK"
        if pal_transito > 0 or pal_transito2 > 0:
            msg += f" (🚢 {pal_transito + pal_transito2} Pal. en tránsito)"
        if incremento > 0:
            msg += f" | +{incremento} Pal. extra"
        return pal_int, pal_merca, pal_txt, pal_avitrans, pal_transito, pal_transito2, seg_pal, cdm_pal, msg, "#155724"

    df[['Pal_Interno', 'Pal_Merca', 'Pal_TXT', 'Pal_Avitrans', 'Pal_Transito', 'Pal_Transito2', 'Seg_pal', 'CDM_pal', 'Estado', 'Color']] = df.apply(
        lambda row: pd.Series(calcular_alerta(row)), axis=1
    )

    # --- Días de cobertura (stock total / CDM) ---
    df['Dias_stock'] = df.apply(
        lambda r: round(r['Pal_Interno'] / r['CDM_pal'])
        if r['CDM_pal'] > 0 else 999, axis=1
    )

    # --- Variación vs semana anterior (desde SQLite) ---
    df['Var_semana'] = 0
    try:
        con = sqlite3.connect(DB_PATH)
        fechas = pd.read_sql("SELECT DISTINCT fecha FROM snapshots ORDER BY fecha DESC LIMIT 8", con)
        fechas_list = fechas['fecha'].tolist()
        if len(fechas_list) >= 7:
            fecha_semana = fechas_list[6]
            stock_ant = pd.read_sql(
                f"SELECT referencia, stock_interno, unidades_palet FROM snapshots WHERE fecha='{fecha_semana}'", con
            )
            stock_ant['u_p'] = stock_ant['unidades_palet'].clip(lower=1)
            stock_ant['pal_ant'] = (stock_ant['stock_interno'] / stock_ant['u_p']).round()
            stock_ant = stock_ant[['referencia', 'pal_ant']].rename(columns={'referencia': 'Referencia'})
            df = df.merge(stock_ant, on='Referencia', how='left')
            df['Var_semana'] = (df['Pal_Interno'] - df['pal_ant'].fillna(df['Pal_Interno'])).round().astype(int)
            df = df.drop(columns=['pal_ant'])
        con.close()
    except Exception:
        pass

    # --- Filtros rápidos ---
    col1, col2 = st.columns(2)
    with col1:
        filtro = st.selectbox("Filtrar por estado:", ["Todos", "🔴 Solo alertas", "🟢 Solo OK"])
    with col2:
        buscar = st.text_input("Buscar referencia:")

    vista = df.copy()
    if filtro == "🔴 Solo alertas":
        vista = vista[vista['Estado'].str.startswith("🔴")]
    elif filtro == "🟢 Solo OK":
        vista = vista[vista['Estado'] == "🟢 OK"]
    if buscar:
        vista = vista[vista['Referencia'].str.contains(buscar, case=False, na=False)]

    # Propagar columnas calculadas a vista
    vista['Dias_stock'] = vista.index.map(df['Dias_stock'])
    vista['Var_semana'] = vista.index.map(df['Var_semana'])

    # --- Métricas resumen ---
    total_alertas  = (df['Estado'].str.startswith("🔴")).sum()
    rotura_critica = (df['Dias_stock'] < df['Lead_time']).sum()

    # Totales por almacén en palets
    total_pal_int = int(df['Pal_Interno'].sum())
    total_pal_txt = int(df['Pal_TXT'].sum())

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Total referencias", len(df))
    m2.metric("🔴 Alertas", total_alertas)
    m3.metric("🟢 OK", len(df) - total_alertas)
    m4.metric("⚠️ Stock < Lead time", rotura_critica)
    m5.metric("🏭 Pal. Interno", total_pal_int)
    m6.metric("📦 Pal. TXT", total_pal_txt)

    # --- Extraer pedido como columna numérica ---
    def extraer_pedido(estado):
        import re
        m = re.search(r'COMPRAR: (\d+)', estado)
        return int(m.group(1)) if m else 0

    df['Pedido_pal'] = df['Estado'].apply(extraer_pedido)
    vista['Pedido_pal'] = vista['Estado'].apply(extraer_pedido)

    # --- Tabla coloreada ---
    cols_mostrar = [
        'Referencia', 'Descripcion', 'Unidades_palet',
        'Seg_pal', 'CDM_pal', 'Dias_stock', 'Var_semana',
        'Pal_Interno', 'Pal_Merca', 'Pal_TXT', 'Pal_Avitrans', 'Pal_Transito', 'Pal_Transito2',
        'Pedido_pal', 'Estado'
    ]
    cols_mostrar = [c for c in cols_mostrar if c in vista.columns]

    def colorear(row):
        color = vista.loc[row.name, 'Color']
        return [f'background-color: {color}; color: white'] * len(row)

    st.dataframe(
        vista[cols_mostrar].style.apply(colorear, axis=1),
        use_container_width=True,
        height=500
    )

    # --- Detalle de productos asociados ---
    if st.session_state.df_materiales is not None:
        st.divider()
        ref_sel = st.selectbox(
            "🔍 Ver productos que consumen esta bandeja:",
            [""] + sorted(vista['Referencia'].tolist()),
            key="ref_detail"
        )
        if ref_sel:
            mat_det = st.session_state.df_materiales.copy()
            mat_det['Codigo'] = mat_det['Codigo'].astype(str).str.strip().str.upper()
            productos = mat_det[mat_det['Codigo'] == ref_sel][['Referencia', 'Descripcion']].drop_duplicates()
            with st.expander(f"📋 Productos que usan {ref_sel} ({len(productos)} referencias)", expanded=True):
                if productos.empty:
                    st.info("No hay productos asociados a esta referencia.")
                else:
                    st.dataframe(productos.reset_index(drop=True), use_container_width=True)

    # --- Exportar a Excel ---
    import io
    output = io.BytesIO()
    export_df = vista[cols_mostrar].copy()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Dashboard')
        ws = writer.sheets['Dashboard']
        # Colores en Excel
        from openpyxl.styles import PatternFill, Font
        color_map = {'#721c24': 'FF721C24', '#856404': 'FF856404', '#155724': 'FF155724'}
        for i, row_idx in enumerate(vista.index, start=2):
            hex_color = color_map.get(vista.loc[row_idx, 'Color'] if row_idx in vista.index else '', 'FF155724')
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            font = Font(color='FFFFFFFF')
            for col_idx in range(1, len(cols_mostrar) + 1):
                cell = ws.cell(row=i, column=col_idx)
                cell.fill = fill
                cell.font = font
    output.seek(0)
    st.download_button(
        label="📥 Exportar a Excel",
        data=output,
        file_name="dashboard_compras.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ══════════════════════════════════════════════
# MÓDULO 3: TRÁNSITO
# ══════════════════════════════════════════════
elif menu == "🚢 Tránsito":
    st.header("🚢 Mercancía en Camino")

    tab1, tab2 = st.tabs(["🌙 Tránsito 1 (noche)", "🌅 Tránsito 2 (mañana)"])

    for tab, key, label in [
        (tab1, 'df_transito',  "Tránsito 1 (noche)"),
        (tab2, 'df_transito2', "Tránsito 2 (mañana)")
    ]:
        with tab:
            col_a, col_b = st.columns(2)
            with col_a:
                st.subheader("Añadir manual")
                with st.form(f"add_{key}", clear_on_submit=True):
                    ref_t  = st.text_input("Referencia", key=f"ref_{key}")
                    cant_t = st.number_input("Cantidad", min_value=0.0, step=1.0, key=f"cant_{key}")
                    if st.form_submit_button("➕ Añadir"):
                        if ref_t.strip():
                            nueva = pd.DataFrame([{'Referencia': ref_t.strip(), 'Cantidad': cant_t}])
                            st.session_state[key] = pd.concat(
                                [st.session_state[key], nueva], ignore_index=True
                            )
                            coleccion_t = 'etiquetas' if key == 'df_transito_etq' else 'bandejas'
                            df_a_firebase(st.session_state[key], coleccion_t, key)
                            st.rerun()
                        else:
                            st.warning("Introduce una referencia.")
            with col_b:
                st.subheader("Cargar desde Excel")
                f_t = st.file_uploader(f"Excel de {label} (.xlsx)", type="xlsx", key=f"file_{key}")
                if f_t and st.button(f"📥 Cargar Excel", key=f"btn_{key}"):
                    df_t = pd.read_excel(f_t)
                    df_t = normalizar_columnas(df_t)
                    if columnas_faltantes(df_t, ['Referencia', 'Cantidad'], label) == []:
                        df_t['Referencia'] = df_t['Referencia'].astype(str).str.strip()
                        df_t['Cantidad']   = pd.to_numeric(df_t['Cantidad'], errors='coerce').fillna(0)
                        st.session_state[key] = df_t[['Referencia', 'Cantidad']]
                        coleccion_t = 'etiquetas' if key == 'df_transito_etq' else 'bandejas'
                        df_a_firebase(df_t[['Referencia', 'Cantidad']], coleccion_t, key)
                        st.success(f"✅ {label} cargado.")
                        st.rerun()
            st.subheader(f"Listado actual — {label}")
            if st.session_state[key].empty:
                st.info("No hay mercancía en tránsito registrada.")
            else:
                st.dataframe(st.session_state[key], use_container_width=True)
                if st.button(f"🗑️ Limpiar {label}", key=f"clear_{key}"):
                    st.session_state[key] = pd.DataFrame(columns=['Referencia', 'Cantidad'])
                    coleccion_t = 'etiquetas' if key == 'df_transito_etq' else 'bandejas'
                    firebase_borrar_df(coleccion_t, key)
                    st.rerun()

# ══════════════════════════════════════════════
# MÓDULO 4: ANÁLISIS
# ══════════════════════════════════════════════
elif menu == "📈 Análisis":
    st.header("📈 Análisis de Consumo")

    if st.session_state.df_final is None or st.session_state.df_consumos is None:
        st.warning("⚠️ Primero carga y sincroniza los archivos.")
        st.stop()

    df   = st.session_state.df_final.copy()
    cons = st.session_state.df_consumos.copy()

    # Tránsito
    t = (
        st.session_state.df_transito
        .groupby('Referencia')['Cantidad'].sum()
        .reset_index().rename(columns={'Cantidad': 'En_transito'})
    )
    df = df.merge(t, on='Referencia', how='left').fillna(0)

    import plotly.express as px

    # ── MÉTRICAS GLOBALES ──────────────────────────────────────
    st.subheader("📦 Totales por Almacén (Palets)")

    # Recalcular palets por almacén
    df['u_p'] = df['Unidades_palet'].clip(lower=1)
    total_int   = (df['Stock_interno'] / df['u_p']).round().sum()
    total_merca = (df['Stock_merca']   / df['u_p']).round().sum()
    total_txt   = (df['Stock_txt']     / df['u_p']).round().sum()
    total_trans = (df['En_transito']   / df['u_p']).round().sum()
    cdm_total   = df['Cdm'].sum()

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("🏭 Interno (AL6+SGA)", f"{int(total_int)} Pal.")
    c2.metric("🏪 Merca", f"{int(total_merca)} Pal.")
    c3.metric("📦 TXT", f"{int(total_txt)} Pal.")
    c4.metric("🚢 En Tránsito", f"{int(total_trans)} Pal.")
    c5.metric("📊 CDM Total/día", f"{cdm_total:.1f} Pal.")

    st.divider()

    # ── REFERENCIAS POR ENCIMA / DEBAJO DEL CDM ───────────────
    st.subheader("🎯 Consumo vs CDM por Referencia")

    # Consumo real último día con movimiento por referencia
    cons['Fecha'] = cons['Fecha'].dt.normalize()
    ultimo_dia = cons.groupby('Referencia')['Fecha'].max().reset_index().rename(columns={'Fecha': 'UltFecha'})
    cons2 = cons.merge(ultimo_dia, on='Referencia')
    cons_ultimo = (
        cons2[cons2['Fecha'] == cons2['UltFecha']]
        .groupby('Referencia')['Cantidad'].sum()
        .reset_index().rename(columns={'Cantidad': 'Consumo_ayer'})
    )
    cons_ultimo['Consumo_ayer'] = cons_ultimo['Consumo_ayer'].abs()

    df_anal = df[['Referencia', 'Descripcion', 'Cdm', 'Unidades_palet']].merge(
        cons_ultimo, on='Referencia', how='left'
    ).fillna(0)
    df_anal['u_p'] = df_anal['Unidades_palet'].clip(lower=1)
    df_anal['Consumo_ayer_pal'] = (df_anal['Consumo_ayer'] / df_anal['u_p']).round(1)
    df_anal['CDM_pal']          = df_anal['Cdm'].apply(lambda x: round(x, 1))
    df_anal['Desviacion']       = df_anal['Consumo_ayer_pal'] - df_anal['CDM_pal']
    df_anal['Estado_CDM']       = df_anal['Desviacion'].apply(
        lambda d: '🔴 Por encima' if d > 0 else ('🟢 Por debajo' if d < 0 else '⚪ Igual')
    )

    filtro_cdm = st.selectbox("Filtrar:", ["Todos", "🔴 Por encima", "🟢 Por debajo"])
    df_vis = df_anal.copy()
    if filtro_cdm != "Todos":
        df_vis = df_vis[df_vis['Estado_CDM'] == filtro_cdm]

    st.dataframe(
        df_vis[['Referencia', 'Descripcion', 'CDM_pal', 'Consumo_ayer_pal', 'Desviacion', 'Estado_CDM']]
        .sort_values('Desviacion', ascending=False),
        use_container_width=True, height=300
    )

    st.divider()

    # ── GRÁFICA: MOVIMIENTOS POR DÍA (semana) ─────────────────
    st.subheader("📅 Movimientos por Día (última semana)")

    cons['DiaSemana'] = cons['Fecha'].dt.day_name()
    cons['Semana']    = cons['Fecha'].dt.isocalendar().week.astype(int)

    ultima_semana = cons['Semana'].max()
    cons_semana   = cons[cons['Semana'] == ultima_semana].copy()

    mov_dia = (
        cons_semana.groupby('Fecha')['Cantidad']
        .apply(lambda x: x.abs().sum())
        .reset_index()
    )
    mov_dia.columns = ['Fecha', 'Unidades']
    mov_dia['% del total'] = (mov_dia['Unidades'] / mov_dia['Unidades'].sum() * 100).round(1)
    mov_dia['Fecha_str']   = mov_dia['Fecha'].dt.strftime('%a %d/%m')

    fig1 = px.bar(
        mov_dia, x='Fecha_str', y='% del total',
        text='% del total', color='% del total',
        color_continuous_scale='RdYlGn_r',
        labels={'Fecha_str': 'Día', '% del total': '% del total semanal'},
        title='% de movimientos por día sobre el total de la semana'
    )
    fig1.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
    fig1.update_layout(showlegend=False, coloraxis_showscale=False)
    st.plotly_chart(fig1, use_container_width=True)

    st.divider()

    # ── GRÁFICA: MOVIMIENTOS POR REFERENCIA ───────────────────
    st.subheader("📦 Movimientos por Referencia (última semana)")

    ref_sel = st.multiselect(
        "Selecciona referencias (vacío = top 10):",
        options=sorted(cons_semana['Referencia'].unique())
    )

    mov_ref = (
        cons_semana.groupby('Referencia')['Cantidad']
        .apply(lambda x: x.abs().sum())
        .reset_index()
    )
    mov_ref.columns = ['Referencia', 'Unidades']
    mov_ref['% del total'] = (mov_ref['Unidades'] / mov_ref['Unidades'].sum() * 100).round(1)

    if ref_sel:
        mov_ref_vis = mov_ref[mov_ref['Referencia'].isin(ref_sel)]
    else:
        mov_ref_vis = mov_ref.nlargest(10, 'Unidades')

    fig2 = px.bar(
        mov_ref_vis.sort_values('Unidades', ascending=True),
        x='Unidades', y='Referencia', orientation='h',
        text='% del total', color='Unidades',
        color_continuous_scale='Blues',
        title='Unidades movidas por referencia (última semana)'
    )
    fig2.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
    fig2.update_layout(showlegend=False, coloraxis_showscale=False)
    st.plotly_chart(fig2, use_container_width=True)

# ══════════════════════════════════════════════
# MÓDULO 5: AGENTE IA
# ══════════════════════════════════════════════
elif menu == "🤖 Agente IA":
    st.header("🤖 Agente de Aprovisionamiento")

    if not GROQ_API_KEY and not ANTHROPIC_API_KEY:
        st.error("❌ No se encontró ninguna API key en el archivo .env")
        st.stop()

    # ── CONTEXTO COMPLETO ─────────────────────────────────────
    def obtener_contexto(pregunta=""):
        """Devuelve solo el contexto relevante según la pregunta."""
        p = pregunta.lower()
        lineas = []
        hoy = pd.Timestamp.now().normalize()
        lineas.append(f"FECHA: {hoy.strftime('%Y-%m-%d')} ({hoy.strftime('%A')})")

        # Detectar qué secciones son relevantes
        habla_etiquetas  = any(w in p for w in ['etiqueta', 'etiqu', 'label', 'stiker', 'stikers'])
        habla_bandejas   = any(w in p for w in ['bandeja', 'band', 'stock interno', 'pal', 'palets'])
        habla_historico  = any(w in p for w in ['ayer', 'semana', 'dia', 'consumo', 'historico', 'tendencia', 'evolucion'])
        habla_pedidos    = any(w in p for w in ['pedido', 'pedir', 'comprar', 'urgente', 'alerta', 'falta'])
        habla_plan       = any(w in p for w in ['planif', 'fabricar', 'produccion', 'hoy'])
        habla_obsoletos  = any(w in p for w in ['obsoleto', 'sin venta', 'muerto', 'parado'])

        # Si es pregunta general o resumen, incluir todo compacto
        es_general = not any([habla_etiquetas, habla_bandejas, habla_historico,
                               habla_pedidos, habla_plan, habla_obsoletos])
        if es_general:
            habla_bandejas = habla_etiquetas = habla_pedidos = True

        # BANDEJAS
        if habla_bandejas or habla_pedidos:
            if st.session_state.df_final is not None:
                df = st.session_state.df_final.copy()
                t1 = st.session_state.df_transito.groupby('Referencia')['Cantidad'].sum().reset_index().rename(columns={'Cantidad':'T1'}) if not st.session_state.df_transito.empty else pd.DataFrame(columns=['Referencia','T1'])
                t2 = st.session_state.df_transito2.groupby('Referencia')['Cantidad'].sum().reset_index().rename(columns={'Cantidad':'T2'}) if not st.session_state.df_transito2.empty else pd.DataFrame(columns=['Referencia','T2'])
                df = df.merge(t1, on='Referencia', how='left').merge(t2, on='Referencia', how='left')
                df['T1'] = df['T1'].fillna(0)
                df['T2'] = df['T2'].fillna(0)
                lineas.append("\n=== BANDEJAS ===")
                ar, aa, av = 0, 0, 0
                filas = []
                for _, r in df.iterrows():
                    u_p = max(r.get('Unidades_palet',1),1)
                    cdm = r.get('Cdm',0)
                    lead = r.get('Lead_time',0)
                    seg = r.get('Stock_seguridad',0)
                    pal_int = round(r.get('Stock_interno',0)/u_p)
                    pal_t = round((r.get('T1',0)+r.get('T2',0))/u_p)
                    stock_final = pal_int + pal_t - (cdm*lead)
                    dias_cob = round(pal_int/cdm) if cdm>0 else 999
                    if stock_final < seg:
                        pedido = math.ceil(seg-stock_final)
                        est = f"ROJO COMPRAR {pedido}pal" if dias_cob<lead else f"AMARILLO COMPRAR {pedido}pal"
                        if dias_cob<lead: ar+=1
                        else: aa+=1
                    else:
                        est = "VERDE"; av+=1
                    filas.append(f"  {r['Referencia']}|{str(r.get('Descripcion',''))[:18]}|int:{pal_int}|cdm:{round(cdm,1)}|cob:{dias_cob}d|{est}")
                # Si pregunta sobre pedidos, solo mostrar alertas
                if habla_pedidos and not es_general:
                    filas = [f for f in filas if 'ROJO' in f or 'AMARILLO' in f]
                lineas.extend(filas[:30])  # max 30 filas
                lineas.append(f"RESUMEN: rojo={ar} amarillo={aa} verde={av}")

        # ETIQUETAS
        if habla_etiquetas or habla_pedidos:
            if st.session_state.df_etiquetas_final is not None:
                df_etq = st.session_state.df_etiquetas_final.copy()
                UMBRAL = 10000
                lineas.append("\n=== ETIQUETAS ===")
                ae_r, ae_a, ae_v = 0, 0, 0
                filas_etq = []
                for _, r in df_etq.iterrows():
                    cons = r.get('Consumo_mes',0)
                    stk = int(r.get('Stock_interno',0))+int(r.get('Stock_merca',0))+int(r.get('Stock_txt',0))
                    objetivo = cons*3 if cons<UMBRAL else cons
                    rot = 'baja' if cons<UMBRAL else 'alta'
                    if stk < objetivo:
                        pedido = math.ceil(objetivo-stk+cons/2)
                        est = f"ROJO COMPRAR {pedido}" if stk<cons/2 else f"AMARILLO COMPRAR {pedido}"
                        if stk<cons/2: ae_r+=1
                        else: ae_a+=1
                    else:
                        est = "VERDE"; ae_v+=1
                    filas_etq.append(f"  {r['Referencia']}|{str(r.get('Descripcion',''))[:18]}|stk:{stk}|cons_mes:{int(cons)}|rot:{rot}|{est}")
                if habla_pedidos and not es_general:
                    filas_etq = [f for f in filas_etq if 'ROJO' in f or 'AMARILLO' in f]
                lineas.extend(filas_etq[:30])
                lineas.append(f"RESUMEN ETQ: rojo={ae_r} amarillo={ae_a} verde={ae_v}")

        # HISTORICO Y CONSUMOS
        if habla_historico:
            try:
                con_db = sqlite3.connect(DB_PATH)
                fechas = pd.read_sql("SELECT DISTINCT fecha FROM snapshots ORDER BY fecha DESC LIMIT 7", con_db)
                fechas_list = fechas['fecha'].tolist()
                if len(fechas_list) >= 2:
                    hist = pd.read_sql(f"SELECT referencia, fecha, stock_interno, unidades_palet FROM snapshots WHERE fecha IN ({','.join([repr(f) for f in fechas_list])})", con_db)
                    hist['u_p'] = hist['unidades_palet'].clip(lower=1)
                    hist['pal'] = (hist['stock_interno']/hist['u_p']).round().astype(int)
                    pivot = hist.pivot_table(index='referencia', columns='fecha', values='pal', aggfunc='sum').fillna(0)
                    pivot = pivot.reindex(columns=sorted(pivot.columns, reverse=True))
                    lineas.append(f"\n=== HISTORICO ({len(fechas_list)} dias) ===")
                    lineas.append("ref | " + " | ".join([f[-5:] for f in sorted(fechas_list, reverse=True)]))
                    for ref, row in list(pivot.iterrows())[:20]:
                        vals = [str(int(v)) for v in row.values]
                        lineas.append(f"  {ref}: {' | '.join(vals)}")
                con_db.close()
            except Exception as e:
                lineas.append(f"[Sin historico: {e}]")

            if st.session_state.df_consumos is not None:
                try:
                    cons = st.session_state.df_consumos.copy()
                    cons['Fecha'] = pd.to_datetime(cons['Fecha'], errors='coerce').dt.normalize()
                    cons['Cantidad'] = cons['Cantidad'].abs()
                    cons['DiaSemana'] = cons['Fecha'].dt.day_name()
                    cons['Semana'] = cons['Fecha'].dt.isocalendar().week.astype(int)
                    total_g = cons['Cantidad'].sum()
                    if total_g > 0:
                        pct = cons.groupby('DiaSemana')['Cantidad'].sum()/total_g*100
                        lineas.append("\n=== % POR DIA SEMANA ===")
                        for dia in ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']:
                            if dia in pct.index:
                                lineas.append(f"  {dia}: {round(pct[dia],1)}%")
                    semana_act = cons['Semana'].max()
                    cs = cons[cons['Semana']==semana_act]['Cantidad'].sum()
                    ca = cons[cons['Semana']==semana_act-1]['Cantidad'].sum()
                    if ca > 0:
                        var = round((cs-ca)/ca*100,1)
                        lineas.append(f"\n=== COMPARATIVA SEMANAL ===")
                        lineas.append(f"  Semana {semana_act}: {int(cs)} ud  vs  Semana {semana_act-1}: {int(ca)} ud  Variacion: {'+' if var>0 else ''}{var}%")
                except Exception:
                    pass

        # PEDIDOS Y PLANIFICACION
        if habla_pedidos or habla_plan or es_general:
            if st.session_state.df_pedidos is not None and not st.session_state.df_pedidos.empty:
                pend = st.session_state.df_pedidos[st.session_state.df_pedidos['Fecha_entrega'] >= hoy]
                if not pend.empty:
                    lineas.append("\n=== PEDIDOS PENDIENTES ===")
                    for _, r in pend.iterrows():
                        lineas.append(f"  {r['Referencia']}: {r['Cantidad']} ud entrega {str(r['Fecha_entrega'])[:10]}")

            if habla_plan and st.session_state.df_planificacion is not None:
                lineas.append("\n=== PLANIFICACION HOY ===")
                for _, r in st.session_state.df_planificacion.iterrows():
                    lineas.append(f"  {r['Codigo']}: {r['Apro']} ud")

        return '\n'.join(lineas)

    contexto_str = obtener_contexto(st.session_state.chat_history[-1]["content"] if st.session_state.chat_history and isinstance(st.session_state.chat_history[-1].get("content"), str) else "")

    # ── FUNCIONES API ──────────────────────────────────────────
    def llamar_groq(messages, ctx):
        from groq import Groq
        client = Groq(api_key=GROQ_API_KEY)
        response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            temperature=0.3,
            max_tokens=1024,
            messages=[
                {"role": "system", "content": f"""Eres un experto en aprovisionamiento y logística.
Tienes acceso a los datos reales del sistema:
{ctx}
Responde en español, de forma concisa y práctica. Usa los datos reales del contexto."""},
                *messages
            ]
        )
        return response.choices[0].message.content

    def llamar_anthropic(messages, ctx):
        import urllib.request
        # Asegurar que los mensajes alternan correctamente user/assistant
        ant_messages = []
        for m in messages:
            if isinstance(m.get("content"), str):
                ant_messages.append({"role": m["role"], "content": m["content"]})
        # Sonnet 4.6 requiere que empiece con user
        if ant_messages and ant_messages[0]["role"] != "user":
            ant_messages = ant_messages[1:]
        payload = {
            "model": "claude-sonnet-4-6",
            "max_tokens": 2048,
            "system": f"""Eres Claude, un asistente experto en gestión de aprovisionamiento, logística y análisis de inventario.
Tienes acceso completo a los datos reales del sistema de compras del usuario:

{ctx}

Puedes responder sobre:
- Estado actual del stock de bandejas y etiquetas
- Alertas de compra y referencias críticas
- Consumos históricos, tendencias y patrones por día de semana
- Cálculos del punto de pedido, CDM, lead time y stock de seguridad
- Planificación de producción y cotejo de materiales
- Pedidos pendientes y tránsitos
- Obsoletos y referencias sin ventas
- Cualquier análisis o pregunta sobre el negocio de aprovisionamiento

Responde siempre en español. Sé preciso, analítico y práctico. 
Cuando des recomendaciones, basa las en los datos reales del contexto.
Si detectas patrones o anomalías en los datos, menciónalos proactivamente.""",
            "messages": ant_messages
        }
        data = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=data,
            headers={'x-api-key': ANTHROPIC_API_KEY, 'anthropic-version': '2023-06-01', 'Content-Type': 'application/json'}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
            return result['content'][0]['text']

    # ── INFO ──────────────────────────────────────────────────
    col_i1, col_i2 = st.columns(2)
    col_i1.info(f"🔑 Claude: {'✅ Conectado' if ANTHROPIC_API_KEY else '❌ Sin key'}")
    col_i2.info(f"⚡ Groq: {'✅ Conectado' if GROQ_API_KEY else '❌ Sin key'}")

    # ── PREGUNTAS RÁPIDAS ─────────────────────────────────────
    st.subheader("💡 Preguntas rápidas")
    preguntas_rapidas = [
        "¿Qué referencias necesito pedir urgentemente?",
        "¿Qué etiquetas están por debajo del consumo mensual?",
        "¿Hay materiales obsoletos con movimientos recientes?",
        "Dame un resumen del estado general del inventario",
        "¿Qué referencias de planificación tienen stock insuficiente?"
    ]
    cols_p = st.columns(len(preguntas_rapidas))
    for i, (col, preg) in enumerate(zip(cols_p, preguntas_rapidas)):
        if col.button(preg, key=f"quick_{i}", use_container_width=True):
            st.session_state.chat_history.append({"role": "user", "content": preg})
            st.rerun()

    st.divider()

    # ── HISTORIAL ─────────────────────────────────────────────
    for msg in st.session_state.chat_history:
        with st.chat_message(msg["role"]):
            if msg["role"] == "assistant" and isinstance(msg["content"], dict):
                col_claude, col_groq = st.columns(2)
                with col_claude:
                    st.markdown("**🤖 Claude**")
                    st.write(msg["content"].get("claude", "Sin respuesta"))
                with col_groq:
                    st.markdown("**⚡ Groq**")
                    st.write(msg["content"].get("groq", "Sin respuesta"))
            else:
                st.write(msg["content"])

    # ── PROCESAR PREGUNTA ─────────────────────────────────────
    if st.session_state.chat_history and st.session_state.chat_history[-1]["role"] == "user":
        ultima_pregunta = st.session_state.chat_history[-1]["content"]
        contexto_completo = obtener_contexto(ultima_pregunta)  # mismo contexto dinámico para Claude
        contexto_dinamico = obtener_contexto(ultima_pregunta)  # contexto reducido para Groq
        messages_para_api = [m for m in st.session_state.chat_history if isinstance(m["content"], str)]

        col_claude, col_groq = st.columns(2)
        resp_claude = None
        resp_groq   = None

        with col_claude:
            st.markdown("**🤖 Claude (contexto completo)**")
            if ANTHROPIC_API_KEY:
                with st.spinner("Claude analizando..."):
                    try:
                        resp_claude = llamar_anthropic(messages_para_api, contexto_completo)
                        st.write(resp_claude)
                    except Exception as e:
                        err_str = str(e)
                        if 'credit' in err_str.lower() or 'balance' in err_str.lower() or 'billing' in err_str.lower():
                            resp_claude = "💳 Sin saldo en Claude. Recarga créditos en console.anthropic.com"
                            st.warning(resp_claude)
                        else:
                            resp_claude = f"Error: {e}"
                            st.error(resp_claude)
            else:
                resp_claude = "❌ Sin API key de Anthropic"
                st.warning(resp_claude)

        with col_groq:
            st.markdown("**⚡ Groq (respuesta rápida)**")
            if GROQ_API_KEY:
                with st.spinner("Groq pensando..."):
                    try:
                        resp_groq = llamar_groq(messages_para_api, contexto_dinamico)
                        st.write(resp_groq)
                    except Exception as e:
                        resp_groq = f"Error: {e}"
                        st.error(resp_groq)
            else:
                resp_groq = "❌ Sin API key de Groq"
                st.warning(resp_groq)

        st.session_state.chat_history.append({
            "role": "assistant",
            "content": {"claude": resp_claude, "groq": resp_groq}
        })

    # ── INPUT ─────────────────────────────────────────────────
    if prompt := st.chat_input("Pregunta algo sobre tu inventario..."):
        st.session_state.chat_history.append({"role": "user", "content": prompt})
        st.rerun()

    if st.session_state.chat_history:
        if st.button("🗑️ Limpiar conversación"):
            st.session_state.chat_history = []
            st.rerun()

# MÓDULO 6: MATERIALES ASOCIADOS
# ══════════════════════════════════════════════
elif menu == "🔗 Materiales":
    st.header("🔗 Materiales Asociados")

    # --- Carga del Excel ---
    f_mat = st.file_uploader("Cargar Excel de Materiales Asociados (.xlsx)", type="xlsx")
    if f_mat:
        df_mat = leer_excel(f_mat, "Materiales")
        if df_mat is not None:
            df_mat = normalizar_columnas(df_mat)
            COL_MAT = ['Gen', 'Referencia', 'Descripcion', 'Codigo', 'Descripcion_material']
            faltan = columnas_faltantes(df_mat, COL_MAT, "Materiales")
            if not faltan:
                # Limpiar
                df_mat['Codigo']     = df_mat['Codigo'].astype(str).str.strip().str.upper()
                df_mat['Referencia'] = df_mat['Referencia'].astype(str).str.strip()
                # Filtrar filas sin código válido
                df_mat = df_mat[~df_mat['Codigo'].isin(['NULL', 'NAN', '', 'NONE'])]
                st.session_state.df_materiales = df_mat[COL_MAT]
                df_a_firebase(df_mat[COL_MAT], 'materiales', 'df_materiales')
                st.success(f"✅ {len(df_mat)} asociaciones cargadas y respaldadas en Firebase.")

    if st.session_state.df_materiales is None:
        st.info("Sube el Excel de materiales para empezar.")
        st.stop()

    df_mat = st.session_state.df_materiales.copy()

    st.divider()

    # --- Buscador por código de material ---
    st.subheader("🔍 Buscar productos que usan un material")

    codigos_disponibles = sorted([str(x) for x in df_mat['Codigo'].unique()])
    
    col_bus1, col_bus2 = st.columns([1, 2])
    with col_bus1:
        buscar_cod = st.text_input("Escribe el código del material:", placeholder="Ej: C12043")
    with col_bus2:
        # Selector con todos los códigos como alternativa
        sel_cod = st.selectbox("O selecciona de la lista:", [""] + codigos_disponibles)

    # Prioridad: texto escrito > selector
    codigo_final = buscar_cod.strip().upper() if buscar_cod.strip() else sel_cod.strip().upper()

    if codigo_final:
        resultado = df_mat[df_mat['Codigo'] == codigo_final]
        if resultado.empty:
            st.warning(f"No se encontraron productos para el código `{codigo_final}`.")
        else:
            # Nombre del material
            nombre_mat = resultado['Descripcion_material'].iloc[0]
            st.markdown(f"### 📦 `{codigo_final}` — {nombre_mat}")
            st.markdown(f"**{len(resultado)} productos** usan este material:")

            st.dataframe(
                resultado[['Gen', 'Referencia', 'Descripcion']].reset_index(drop=True),
                use_container_width=True,
                height=400
            )

            # Exportar
            import io
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                resultado[['Gen', 'Referencia', 'Descripcion', 'Codigo', 'Descripcion_material']].to_excel(
                    writer, index=False, sheet_name='Materiales'
                )
            output.seek(0)
            st.download_button(
                label="📥 Exportar a Excel",
                data=output,
                file_name=f"materiales_{codigo_final}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    st.divider()

    # --- Búsqueda inversa: dado un producto, ver sus materiales ---
    st.subheader("🔄 Buscar materiales de un producto")

    refs_disponibles = sorted([str(x) for x in df_mat['Referencia'].unique()])
    buscar_ref = st.selectbox("Selecciona una referencia de producto:", [""] + refs_disponibles)

    if buscar_ref:
        resultado_inv = df_mat[df_mat['Referencia'] == buscar_ref]
        nombre_prod = resultado_inv['Descripcion'].iloc[0]
        st.markdown(f"### 🏷️ `{buscar_ref}` — {nombre_prod}")
        st.markdown(f"**{len(resultado_inv)} materiales** usa este producto:")
        st.dataframe(
            resultado_inv[['Codigo', 'Descripcion_material']].reset_index(drop=True),
            use_container_width=True
        )

# ══════════════════════════════════════════════
# MÓDULO 7: ETIQUETAS
# ══════════════════════════════════════════════
elif menu == "🏷️ Etiquetas":
    st.header("🏷️ Gestión de Etiquetas")

    COL_MAESTRO_ETQ = ['Referencia', 'Descripcion', 'Lead_time', 'Multiplicador', 'Unidades_caja', 'Esetiquetadecaja']
    COL_VENTAS      = ['Referencia', 'Unidades']
    UMBRAL_BAJA_ROT = 10000  # Unidades/mes

    # ── CARGA DE ARCHIVOS ─────────────────────────────────────
    st.subheader("📂 Cargar Archivos")
    col_c1, col_c2, col_c3, col_c4 = st.columns(4)
    with col_c1:
        f_metq = st.file_uploader("1. Maestro Etiquetas (.xlsx)", type="xlsx", key="metq")
    with col_c2:
        f_vent = st.file_uploader("2. Ventas del mes (.xlsx)", type="xlsx", key="vent")
        st.caption("Columnas: Referencia, Unidades (acumulado mensual)")
    with col_c3:
        f_setq = st.file_uploader("3. Stock Actual (.xlsx)", type="xlsx", key="setq")
    with col_c4:
        f_env = st.file_uploader("4. Etiquetas por envase (.xlsx)", type="xlsx", key="fenv")
        st.caption("Columnas: Referencia, Etiquetas_envase")
        if f_env:
            df_env_up = leer_excel(f_env, "Envase")
            if df_env_up is not None:
                df_env_up = normalizar_columnas(df_env_up)
                if not columnas_faltantes(df_env_up, ['Referencia', 'Etiquetas_envase'], "Envase"):
                    df_env_up['Referencia'] = df_env_up['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                    df_env_up['Etiquetas_envase'] = pd.to_numeric(df_env_up['Etiquetas_envase'], errors='coerce').fillna(1).clip(lower=1)
                    st.session_state.df_envase = df_env_up[['Referencia', 'Etiquetas_envase']]
                    df_a_firebase(st.session_state.df_envase, 'etiquetas', 'df_envase')
                    st.success(f"✅ {len(df_env_up)} referencias de envase guardadas.")

    if st.button("🚀 Sincronizar Etiquetas"):
        if not (f_metq and f_vent and f_setq):
            st.error("Sube los 3 archivos.")
            st.stop()

        m_etq  = leer_excel(f_metq,  "Maestro Etiquetas")
        ventas = leer_excel(f_vent,  "Ventas")
        s_etq  = leer_excel(f_setq,  "Stock Etiquetas")
        if m_etq is None or ventas is None or s_etq is None:
            st.stop()

        m_etq  = normalizar_columnas(m_etq)
        ventas = normalizar_columnas(ventas)
        s_etq  = normalizar_columnas(s_etq)

        errores = (
            columnas_faltantes(m_etq,  COL_MAESTRO_ETQ, "Maestro Etiquetas") +
            columnas_faltantes(ventas, COL_VENTAS,       "Ventas") +
            columnas_faltantes(s_etq,  COL_STOCK,        "Stock Etiquetas")
        )
        if errores:
            st.stop()

        for df in [m_etq, ventas, s_etq]:
            df['Referencia'] = df['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)

        for col in ['Lead_time', 'Multiplicador', 'Unidades_caja']:
            m_etq[col] = pd.to_numeric(m_etq[col], errors='coerce').fillna(0)
        m_etq['Esetiquetadecaja'] = m_etq['Esetiquetadecaja'].astype(str).str.strip().str.lower().fillna('')

        # ── STOCK ─────────────────────────────────────────────
        s_etq['Almacen']  = s_etq['Almacen'].astype(str).str.strip()
        s_etq['Cantidad'] = pd.to_numeric(s_etq['Cantidad'], errors='coerce').fillna(0)
        res_stock_etq = (
            s_etq.groupby('Referencia')
             .apply(lambda g: pd.Series({
                 'Stock_interno': g.loc[g['Almacen'].isin(ALMACENES_INT),   'Cantidad'].sum(),
                 'Stock_merca':   g.loc[g['Almacen'].isin(ALMACENES_MERCA), 'Cantidad'].sum(),
                 'Stock_txt':     g.loc[g['Almacen'].isin(ALMACENES_TXT),   'Cantidad'].sum(),
             }))
             .reset_index()
        )

        # ── CONSUMO MENSUAL DESDE VENTAS ───────────────────────
        ventas['Unidades'] = pd.to_numeric(ventas['Unidades'], errors='coerce').fillna(0).abs()
        # Preservar Descripcion si existe
        if 'Descripcion' not in ventas.columns:
            for col in ventas.columns:
                if 'desc' in col.lower():
                    ventas = ventas.rename(columns={col: 'Descripcion'})
                    break

        if st.session_state.df_materiales is None:
            st.error("❌ Primero carga el Excel de Materiales Asociados en el módulo 🔗 Materiales.")
            st.stop()

        mat = st.session_state.df_materiales.copy()
        mat['Referencia'] = mat['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
        mat['Codigo']     = mat['Codigo'].astype(str).str.strip().str.upper()

        # Ventas x materiales → consumo mensual bruto por etiqueta
        ventas_mat = ventas.merge(mat[['Referencia', 'Codigo']], on='Referencia', how='inner')
        ventas_mat = ventas_mat.rename(columns={'Codigo': 'Ref_etiqueta'})

        ventas_mat = ventas_mat.merge(
            m_etq[['Referencia', 'Multiplicador', 'Unidades_caja', 'Esetiquetadecaja']].rename(columns={'Referencia': 'Ref_etiqueta'}),
            on='Ref_etiqueta', how='left'
        )
        ventas_mat['Multiplicador']     = ventas_mat['Multiplicador'].fillna(1).clip(lower=1)
        ventas_mat['Unidades_caja']     = ventas_mat['Unidades_caja'].fillna(0)
        ventas_mat['Esetiquetadecaja']  = ventas_mat['Esetiquetadecaja'].astype(str).str.strip().str.lower().fillna('')

        # Unir datos de envase (etiquetas de caja) si están disponibles
        if st.session_state.df_envase is not None:
            df_env = st.session_state.df_envase.copy()
            ventas_mat = ventas_mat.merge(df_env, left_on='Referencia', right_on='Referencia', how='left')
            ventas_mat['Etiquetas_envase'] = ventas_mat['Etiquetas_envase'].fillna(1).clip(lower=1)
        else:
            ventas_mat['Etiquetas_envase'] = ventas_mat['Unidades_caja'].fillna(1).clip(lower=1)

        ventas_mat['Consumo_mes'] = ventas_mat.apply(
            lambda r: r['Unidades'] / r['Etiquetas_envase'] if r['Esetiquetadecaja'] == 'si'
                      else r['Unidades'] * max(r['Multiplicador'], 1),
            axis=1
        )

        # Consumo mensual total por etiqueta (suma de todos los productos que la usan)
        consumo_mes = (
            ventas_mat.groupby('Ref_etiqueta')['Consumo_mes']
            .sum().reset_index()
            .rename(columns={'Ref_etiqueta': 'Referencia', 'Consumo_mes': 'Consumo_mes'})
        )

        # ── UNIÓN FINAL ────────────────────────────────────────
        final_etq = pd.merge(m_etq, res_stock_etq, on='Referencia', how='left')
        final_etq = pd.merge(final_etq, consumo_mes, on='Referencia', how='left')
        for col in ['Consumo_mes', 'Stock_interno', 'Stock_merca', 'Stock_txt']:
            final_etq[col] = pd.to_numeric(final_etq[col], errors='coerce').fillna(0)

        st.session_state.df_etiquetas_final = final_etq
        st.session_state.df_ventas = ventas
        with st.spinner("Guardando en Firebase..."):
            ok1, e1 = df_a_firebase(final_etq, 'etiquetas', 'df_etiquetas_final')
            ok2, e2 = df_a_firebase(ventas,    'etiquetas', 'df_ventas')
            if ok1 and ok2:
                st.success(f"✅ {len(final_etq)} etiquetas sincronizadas.")
            else:
                st.warning(f"⚠️ Sincronizado localmente pero error en Firebase: {e1 or e2}")
        st.dataframe(final_etq.head(10), use_container_width=True)

    # ── TRÁNSITO ETIQUETAS ────────────────────────────────────
    with st.expander("🚢 Gestionar Tránsito de Etiquetas"):
        col_te1, col_te2 = st.columns(2)
        with col_te1:
            f_tetq = st.file_uploader("Subir Excel de Tránsito (.xlsx)", type="xlsx", key="tetq")
            if f_tetq and st.button("📥 Cargar Tránsito"):
                df_tetq = pd.read_excel(f_tetq)
                df_tetq = normalizar_columnas(df_tetq)
                df_tetq['Referencia'] = df_tetq['Referencia'].astype(str).str.strip().str.upper()
                df_tetq['Cantidad']   = pd.to_numeric(df_tetq['Cantidad'], errors='coerce').fillna(0)
                st.session_state.df_transito_etq = df_tetq[['Referencia', 'Cantidad']]
                df_a_firebase(st.session_state.df_transito_etq, 'etiquetas', 'df_transito_etq')
                st.success("✅ Tránsito de etiquetas cargado.")
                st.rerun()
        with col_te2:
            if not st.session_state.df_transito_etq.empty:
                st.dataframe(st.session_state.df_transito_etq, use_container_width=True)
                if st.button("🗑️ Limpiar tránsito etiquetas"):
                    st.session_state.df_transito_etq = pd.DataFrame(columns=['Referencia', 'Cantidad'])
                    firebase_borrar_df('etiquetas', 'df_transito_etq')
                    st.rerun()

    # ── DASHBOARD ETIQUETAS ────────────────────────────────────
    if st.session_state.df_etiquetas_final is not None:
        st.divider()
        st.subheader("📊 Dashboard Etiquetas")

        df_etq = st.session_state.df_etiquetas_final.copy()

        # Añadir tránsito
        t_etq = (
            st.session_state.df_transito_etq
            .groupby('Referencia')['Cantidad'].sum()
            .reset_index().rename(columns={'Cantidad': 'En_transito'})
        )
        df_etq = df_etq.merge(t_etq, on='Referencia', how='left')
        df_etq['En_transito'] = df_etq['En_transito'].fillna(0)

        def alerta_etq(row):
            consumo_mes  = row.get('Consumo_mes', 0)
            stock_total  = row['Stock_interno'] + row['Stock_merca'] + row['Stock_txt'] + row['En_transito']
            transito_ud  = int(row['En_transito'])
            consumo_2sem = consumo_mes / 2  # Lead time = 2 semanas

            baja_rotacion = consumo_mes < UMBRAL_BAJA_ROT

            if baja_rotacion:
                stock_objetivo = consumo_mes * 3  # 3 meses
                label_cdm = f"{int(consumo_mes)} ud/mes (baja rot.)"
            else:
                stock_objetivo = consumo_mes      # 1 mes
                label_cdm = f"{int(consumo_mes)} ud/mes"

            if stock_total < stock_objetivo:
                pedido = math.ceil(stock_objetivo - stock_total + consumo_2sem)
                estado_base = f"COMPRAR: {int(pedido)} Ud."
                # Rojo si stock < 2 semanas, amarillo si no
                if stock_total < consumo_2sem:
                    return int(stock_total), int(row['Stock_interno']), int(row['Stock_merca']), int(row['Stock_txt']), transito_ud, int(consumo_mes), int(pedido), f"🔴 {estado_base}", "#721c24"
                else:
                    return int(stock_total), int(row['Stock_interno']), int(row['Stock_merca']), int(row['Stock_txt']), transito_ud, int(consumo_mes), int(pedido), f"🟡 {estado_base}", "#856404"

            msg = "🟢 OK"
            if transito_ud > 0:
                msg += f" (🚢 {transito_ud} en tránsito)"
            return int(stock_total), int(row['Stock_interno']), int(row['Stock_merca']), int(row['Stock_txt']), transito_ud, int(consumo_mes), 0, msg, "#155724"

        df_etq[['Stock_total', 'Stk_Interno', 'Stk_Merca', 'Stk_TXT', 'Transito_ud', 'CDM_mes', 'Pedido_ud', 'Estado', 'Color']] = df_etq.apply(
            lambda row: pd.Series(alerta_etq(row)), axis=1
        )

        # Filtros
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            filtro_etq = st.selectbox("Filtrar:", ["Todos", "🔴 Solo alertas", "🟡 Amarillo", "🟢 Solo OK"], key="fetq")
        with col_f2:
            buscar_etq = st.text_input("Buscar referencia:", key="betq")

        vista_etq = df_etq.copy()
        if filtro_etq == "🔴 Solo alertas":
            vista_etq = vista_etq[vista_etq['Estado'].str.startswith("🔴")]
        elif filtro_etq == "🟡 Amarillo":
            vista_etq = vista_etq[vista_etq['Estado'].str.startswith("🟡")]
        elif filtro_etq == "🟢 Solo OK":
            vista_etq = vista_etq[vista_etq['Estado'].str.startswith("🟢")]
        if buscar_etq:
            vista_etq = vista_etq[vista_etq['Referencia'].str.contains(buscar_etq, case=False, na=False)]

        # Métricas
        alertas_etq    = (df_etq['Estado'].str.startswith("🔴")).sum()
        amarillas_etq  = (df_etq['Estado'].str.startswith("🟡")).sum()
        total_int_etq  = int(df_etq['Stk_Interno'].sum())
        total_txt_etq  = int(df_etq['Stk_TXT'].sum())
        me1, me2, me3, me4, me5, me6 = st.columns(6)
        me1.metric("Total etiquetas", len(df_etq))
        me2.metric("🔴 Alertas", alertas_etq)
        me3.metric("🟡 Avisos", amarillas_etq)
        me4.metric("🟢 OK", len(df_etq) - alertas_etq - amarillas_etq)
        me5.metric("🏭 Stk Interno", total_int_etq)
        me6.metric("📦 Stk TXT", total_txt_etq)

        cols_etq = ['Referencia', 'Descripcion', 'CDM_mes', 'Stk_Interno', 'Stk_Merca', 'Stk_TXT', 'Transito_ud', 'Pedido_ud', 'Estado']

        def colorear_etq(row):
            color = vista_etq.loc[row.name, 'Color']
            return [f'background-color: {color}; color: white'] * len(row)

        st.dataframe(
            vista_etq[cols_etq].style.apply(colorear_etq, axis=1),
            use_container_width=True, height=500
        )

        # Exportar
        import io
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            vista_etq[cols_etq].to_excel(writer, index=False, sheet_name='Etiquetas')
        output.seek(0)
        st.download_button("📥 Exportar a Excel", output, "dashboard_etiquetas.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════
# MÓDULO 8: PEDIDOS
# ══════════════════════════════════════════════
elif menu == "📋 Pedidos":
    st.header("📋 Histórico de Pedidos")

    # ── Cargar Excel de pedidos ───────────────
    f_ped = st.file_uploader("Subir Excel de Pedidos (.xlsx)", type="xlsx", key="fped")
    if f_ped and st.button("📥 Cargar Pedidos"):
        df_ped = pd.read_excel(f_ped)
        df_ped = normalizar_columnas(df_ped)
        faltan = columnas_faltantes(df_ped, ['Referencia', 'Cantidad', 'Fecha_entrega'], "Pedidos")
        if not faltan:
            df_ped['Referencia']    = df_ped['Referencia'].astype(str).str.strip().str.upper()
            df_ped['Cantidad']      = pd.to_numeric(df_ped['Cantidad'], errors='coerce').fillna(0)
            df_ped['Fecha_entrega'] = pd.to_datetime(df_ped['Fecha_entrega'], errors='coerce')
            # Acumular sobre los existentes
            st.session_state.df_pedidos = pd.concat(
                [st.session_state.df_pedidos, df_ped[['Referencia', 'Cantidad', 'Fecha_entrega']]],
                ignore_index=True
            ).drop_duplicates()
            df_a_firebase(st.session_state.df_pedidos, 'pedidos', 'df_pedidos')
            st.success(f"✅ {len(df_ped)} pedidos cargados y guardados en Firebase.")
            st.rerun()

    st.divider()

    if st.session_state.df_pedidos.empty:
        st.info("No hay pedidos registrados. Sube un Excel con columnas: Referencia, Cantidad, Fecha_entrega.")
        st.stop()

    df_ped = st.session_state.df_pedidos.copy()
    df_ped['Fecha_entrega'] = pd.to_datetime(df_ped['Fecha_entrega'], errors='coerce')

    # ── Filtros ───────────────────────────────
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        buscar_ref = st.text_input("Buscar referencia:")
    with col_f2:
        desde = st.date_input("Desde:", value=df_ped['Fecha_entrega'].min().date() if not df_ped.empty else None)
    with col_f3:
        hasta = st.date_input("Hasta:", value=df_ped['Fecha_entrega'].max().date() if not df_ped.empty else None)

    vista_ped = df_ped.copy()
    if buscar_ref:
        vista_ped = vista_ped[vista_ped['Referencia'].str.contains(buscar_ref, case=False, na=False)]
    if desde:
        vista_ped = vista_ped[vista_ped['Fecha_entrega'] >= pd.Timestamp(desde)]
    if hasta:
        vista_ped = vista_ped[vista_ped['Fecha_entrega'] <= pd.Timestamp(hasta)]

    vista_ped = vista_ped.sort_values('Fecha_entrega', ascending=False)

    # ── Métricas ─────────────────────────────
    hoy = pd.Timestamp.now().normalize()
    pendientes = df_ped[df_ped['Fecha_entrega'] >= hoy]
    mp1, mp2, mp3 = st.columns(3)
    mp1.metric("Total pedidos", len(df_ped))
    mp2.metric("Pendientes de entrega", len(pendientes))
    mp3.metric("Referencias distintas", df_ped['Referencia'].nunique())

    # ── Tabla ─────────────────────────────────
    st.dataframe(vista_ped.reset_index(drop=True), use_container_width=True, height=400)

    # ── Exportar ──────────────────────────────
    import io
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        vista_ped.to_excel(writer, index=False, sheet_name='Pedidos')
    output.seek(0)
    st.download_button("📥 Exportar a Excel", output, "pedidos.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── Borrar todos ──────────────────────────
    if st.button("🗑️ Borrar todos los pedidos"):
        st.session_state.df_pedidos = pd.DataFrame(columns=['Referencia', 'Cantidad', 'Fecha_entrega'])
        firebase_borrar_df('pedidos', 'df_pedidos')
        st.success("Pedidos borrados.")
        st.rerun()

# ══════════════════════════════════════════════
# MÓDULO 9: PLANIFICACIÓN
# ══════════════════════════════════════════════
elif menu == "🔍 Previsión y Obsoletos":
    st.header("🔍 Previsión y Obsoletos")

    f_plan = st.file_uploader("Subir archivo de Planificación (.xlsx)", type="xlsx", key="fplan")

    if f_plan and st.button("📥 Cargar Planificación"):
        df_plan = leer_excel(f_plan, "Planificación")
        if df_plan is not None:
            df_plan = normalizar_columnas(df_plan)
            col_map = {}
            for c in df_plan.columns:
                cl = c.lower()
                if 'cod' in cl: col_map[c] = 'Codigo'
                elif 'desc' in cl: col_map[c] = 'Descripcion'
                elif 'apro' in cl or 'unidad' in cl or 'cant' in cl: col_map[c] = 'Apro'
            df_plan = df_plan.rename(columns=col_map)
            faltan = columnas_faltantes(df_plan, ['Codigo', 'Apro'], "Planificación")
            if not faltan:
                df_plan['Codigo'] = df_plan['Codigo'].astype(str).str.strip().str.upper()
                df_plan['Apro']   = pd.to_numeric(df_plan['Apro'], errors='coerce').fillna(0)
                df_plan = df_plan[df_plan['Apro'] > 0]
                st.session_state.df_planificacion = df_plan
                df_a_firebase(df_plan, 'planificacion', 'df_planificacion')
                st.success(f"✅ {len(df_plan)} materiales cargados.")
                st.rerun()

    if st.session_state.df_planificacion is None:
        st.info("Sube el archivo de planificación para empezar.")
        st.stop()

    df_plan = st.session_state.df_planificacion.copy()

    if st.session_state.df_final is None:
        st.error("❌ Sincroniza primero el módulo de Bandejas.")
        st.stop()

    # Preparar datos
    df_band = st.session_state.df_final.copy()
    df_band['Referencia'] = df_band['Referencia'].astype(str).str.strip().str.upper()
    refs_maestro_band = set(df_band['Referencia'].unique())
    for col in ['Stock_merca', 'Stock_txt', 'Stock_avitrans']:
        if col not in df_band.columns:
            df_band[col] = 0
    # Solo stock interno para planificacion
    df_band['Stock_total_ud'] = df_band['Stock_interno'].fillna(0)

    df_etq_cp = None
    refs_maestro_etq = set()
    if st.session_state.df_etiquetas_final is not None:
        df_etq_cp = st.session_state.df_etiquetas_final.copy()
        df_etq_cp['Referencia'] = df_etq_cp['Referencia'].astype(str).str.strip().str.upper()
        # Solo stock interno para planificacion
        df_etq_cp['Stock_total_ud'] = df_etq_cp['Stock_interno'].fillna(0)
        refs_maestro_etq = set(df_etq_cp['Referencia'].unique())

    import io
    tab1, tab2, tab3 = st.tabs(["📦 Stock Insuficiente", "❓ Sin Maestro", "🗃️ Obsoletos"])

    # ══ TAB 1: STOCK INSUFICIENTE ══════════════════════════════
    with tab1:
        st.subheader("Materiales con stock insuficiente para la planificación")

        # Planificación ya viene desglosada por material directamente
        # Agrupar por Codigo sumando Apro (puede haber varias líneas del mismo material)
        necesidad = df_plan.groupby('Codigo')['Apro'].sum().reset_index()
        necesidad.columns = ['Referencia', 'Necesidad_ud']

        # ── BANDEJAS ──────────────────────────────────────────
        st.markdown("#### 📦 Bandejas (en box)")
        cot_band = df_band[['Referencia', 'Descripcion', 'Stock_total_ud', 'Unidades_palet']].copy()
        cot_band = cot_band.merge(necesidad, on='Referencia', how='left')
        cot_band['Necesidad_ud']   = cot_band['Necesidad_ud'].fillna(0)
        cot_band['Unidades_palet'] = cot_band['Unidades_palet'].fillna(1).clip(lower=1)
        cot_band['Nec_box']        = (cot_band['Necesidad_ud']   / cot_band['Unidades_palet']).apply(math.ceil)
        cot_band['Stock_box']      = (cot_band['Stock_total_ud'] / cot_band['Unidades_palet']).apply(math.floor)
        cot_band['Dif_box']        = cot_band['Stock_box'] - cot_band['Nec_box']
        cot_band['Estado']         = cot_band.apply(
            lambda r: f"🔴 FALTA: {abs(int(r['Dif_box']))} box" if r['Dif_box'] < 0
            else ("⚪ Sin planificar" if r['Necesidad_ud'] == 0 else "🟢 OK"), axis=1
        )
        cot_band['Color'] = cot_band['Dif_box'].apply(lambda d: "#721c24" if d < 0 else "#155724")

        alertas_b = (cot_band['Dif_box'] < 0).sum()
        cb1, cb2, cb3 = st.columns(3)
        cb1.metric("Total bandejas", len(cot_band))
        cb2.metric("🔴 Falta stock", alertas_b)
        cb3.metric("📋 Planificadas hoy", (cot_band['Necesidad_ud'] > 0).sum())

        filtro_band = st.selectbox("Filtrar bandejas:", ["Todas", "🔴 Solo con falta", "📋 Solo planificadas"], key="fb")
        vista_band = cot_band.copy()
        if filtro_band == "🔴 Solo con falta":
            vista_band = vista_band[vista_band['Dif_box'] < 0]
        elif filtro_band == "📋 Solo planificadas":
            vista_band = vista_band[vista_band['Necesidad_ud'] > 0]

        cols_b = ['Referencia', 'Descripcion', 'Nec_box', 'Stock_box', 'Dif_box', 'Estado']
        def color_band(row):
            return [f'background-color: {cot_band.loc[row.name, "Color"]}; color: white'] * len(row)
        st.dataframe(vista_band[cols_b].style.apply(color_band, axis=1), use_container_width=True, height=400)
        out_b = io.BytesIO()
        with pd.ExcelWriter(out_b, engine='openpyxl') as w:
            vista_band[cols_b].to_excel(w, index=False, sheet_name='Bandejas')
        out_b.seek(0)
        st.download_button("📥 Exportar Bandejas", out_b, "cotejo_bandejas.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # ── ETIQUETAS ─────────────────────────────────────────
        if df_etq_cp is not None:
            st.markdown("#### 🏷️ Etiquetas (en unidades)")
            cot_etq = df_etq_cp[['Referencia', 'Descripcion', 'Stock_total_ud']].copy()
            cot_etq = cot_etq.merge(necesidad, on='Referencia', how='left')
            cot_etq['Necesidad_ud'] = cot_etq['Necesidad_ud'].fillna(0)
            cot_etq['Dif_ud']  = cot_etq['Stock_total_ud'] - cot_etq['Necesidad_ud']
            cot_etq['Estado']  = cot_etq.apply(
                lambda r: f"🔴 FALTA: {abs(int(r['Dif_ud']))} ud" if r['Dif_ud'] < 0
                else ("⚪ Sin planificar" if r['Necesidad_ud'] == 0 else "🟢 OK"), axis=1
            )
            cot_etq['Color'] = cot_etq['Dif_ud'].apply(lambda d: "#721c24" if d < 0 else "#155724")

            alertas_e = (cot_etq['Dif_ud'] < 0).sum()
            ce1, ce2, ce3 = st.columns(3)
            ce1.metric("Total etiquetas", len(cot_etq))
            ce2.metric("🔴 Falta stock", alertas_e)
            ce3.metric("📋 Planificadas hoy", (cot_etq['Necesidad_ud'] > 0).sum())

            filtro_etq = st.selectbox("Filtrar etiquetas:", ["Todas", "🔴 Solo con falta", "📋 Solo planificadas"], key="fe")
            vista_etq = cot_etq.copy()
            if filtro_etq == "🔴 Solo con falta":
                vista_etq = vista_etq[vista_etq['Dif_ud'] < 0]
            elif filtro_etq == "📋 Solo planificadas":
                vista_etq = vista_etq[vista_etq['Necesidad_ud'] > 0]

            cols_e = ['Referencia', 'Descripcion', 'Necesidad_ud', 'Stock_total_ud', 'Dif_ud', 'Estado']
            def color_etq_p(row):
                return [f'background-color: {cot_etq.loc[row.name, "Color"]}; color: white'] * len(row)
            st.dataframe(vista_etq[cols_e].style.apply(color_etq_p, axis=1), use_container_width=True, height=400)
            out_e = io.BytesIO()
            with pd.ExcelWriter(out_e, engine='openpyxl') as w:
                vista_etq[cols_e].to_excel(w, index=False, sheet_name='Etiquetas')
            out_e.seek(0)
            st.download_button("📥 Exportar Etiquetas", out_e, "cotejo_etiquetas.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ══ TAB 2: SIN MAESTRO ══════════════════════════════════════
    with tab2:
        st.subheader("Materiales en planificación no configurados en ningún maestro")
        refs_todos_maestros = refs_maestro_band | refs_maestro_etq
        sin_maestro = df_plan[~df_plan['Codigo'].isin(refs_todos_maestros)].copy()
        if sin_maestro.empty:
            st.success("✅ Todos los materiales están configurados en los maestros.")
        else:
            st.warning(f"⚠️ {len(sin_maestro)} materiales sin configurar:")
            cols_sm = [c for c in ['Codigo', 'Descripcion', 'Apro'] if c in sin_maestro.columns]
            st.dataframe(sin_maestro[cols_sm].reset_index(drop=True), use_container_width=True)
            out_sm = io.BytesIO()
            with pd.ExcelWriter(out_sm, engine='openpyxl') as w:
                sin_maestro[cols_sm].to_excel(w, index=False, sheet_name='Sin_maestro')
            out_sm.seek(0)
            st.download_button("📥 Exportar Sin Maestro", out_sm, "sin_maestro.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ══ TAB 3: OBSOLETOS ════════════════════════════════════════
    with tab3:
        st.subheader("Materiales con stock pero sin ventas este mes")
        if st.session_state.df_ventas is None:
            st.warning("⚠️ Carga primero el módulo de Etiquetas con el archivo de ventas.")
        else:
            ventas_refs = st.session_state.df_ventas['Referencia'].astype(str).str.strip().str.upper().str.zfill(6).unique()
            if st.session_state.df_materiales is not None:
                mat_obs = st.session_state.df_materiales.copy()
                mat_obs['Referencia'] = mat_obs['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                mat_obs['Codigo']     = mat_obs['Codigo'].astype(str).str.strip().str.upper()
                mat_con_ventas = set(mat_obs[mat_obs['Referencia'].isin(ventas_refs)]['Codigo'].unique())
            else:
                mat_con_ventas = set()

            # Referencias con movimientos en consumos
            refs_con_movimientos = set()
            if st.session_state.df_consumos is not None:
                cons_obs = st.session_state.df_consumos.copy()
                cons_obs['Referencia'] = cons_obs['Referencia'].astype(str).str.strip().str.upper()
                refs_con_movimientos = set(cons_obs['Referencia'].unique())

            obs_list = []
            obs_band = df_band[(df_band['Stock_total_ud'] > 0) & (~df_band['Referencia'].isin(mat_con_ventas))][['Referencia', 'Descripcion', 'Stock_total_ud']].copy()
            obs_band.columns = ['Referencia', 'Descripcion', 'Stock']
            obs_band['Tipo'] = 'Bandeja'
            obs_list.append(obs_band)

            if df_etq_cp is not None:
                obs_etq = df_etq_cp[(df_etq_cp['Stock_total_ud'] > 0) & (~df_etq_cp['Referencia'].isin(mat_con_ventas))][['Referencia', 'Descripcion', 'Stock_total_ud']].copy()
                obs_etq.columns = ['Referencia', 'Descripcion', 'Stock']
                obs_etq['Tipo'] = 'Etiqueta'
                obs_list.append(obs_etq)

            obsoletos = pd.concat(obs_list, ignore_index=True)
            obsoletos['Movimientos'] = obsoletos['Referencia'].apply(
                lambda r: '🔴 CON MOVIMIENTOS' if r in refs_con_movimientos else '⚪ Sin movimientos'
            )

            if obsoletos.empty:
                st.success("✅ No hay materiales con stock sin ventas.")
            else:
                con_mov = (obsoletos['Movimientos'] == '🔴 CON MOVIMIENTOS').sum()
                oo1, oo2, oo3 = st.columns(3)
                oo1.metric("Total obsoletos", len(obsoletos))
                oo2.metric("🔴 Con movimientos", con_mov)
                oo3.metric("⚪ Sin movimientos", len(obsoletos) - con_mov)

                filtro_obs = st.selectbox("Filtrar:", ["Todos", "🔴 Con movimientos", "⚪ Sin movimientos", "Bandeja", "Etiqueta"], key="fobs")
                vista_obs = obsoletos.copy()
                if filtro_obs == "🔴 Con movimientos":
                    vista_obs = vista_obs[vista_obs['Movimientos'] == '🔴 CON MOVIMIENTOS']
                elif filtro_obs == "⚪ Sin movimientos":
                    vista_obs = vista_obs[vista_obs['Movimientos'] == '⚪ Sin movimientos']
                elif filtro_obs in ["Bandeja", "Etiqueta"]:
                    vista_obs = vista_obs[vista_obs['Tipo'] == filtro_obs]

                def colorear_obs(row):
                    styles = [''] * len(row.index)
                    try:
                        idx_mov = list(row.index).index('Movimientos')
                        if row['Movimientos'] == '🔴 CON MOVIMIENTOS':
                            styles[idx_mov] = 'background-color: #721c24; color: white'
                    except Exception:
                        pass
                    return styles

                st.dataframe(vista_obs.reset_index(drop=True).style.apply(colorear_obs, axis=1), use_container_width=True)

                out_obs = io.BytesIO()
                with pd.ExcelWriter(out_obs, engine='openpyxl') as w:
                    vista_obs.to_excel(w, index=False, sheet_name='Obsoletos')
                out_obs.seek(0)
                st.download_button("📥 Exportar Obsoletos", out_obs, "obsoletos.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════
# MÓDULO 10: PRODUCTO TERMINADO
# ══════════════════════════════════════════════
elif menu == "🏪 Producto Terminado":
    st.header("🏪 Producto Terminado")

    DIAS_MES = 22

    # ── Carga de archivos ─────────────────────
    st.subheader("📂 Cargar Archivos")
    col1, col2 = st.columns(2)
    with col1:
        f_spt = st.file_uploader("1. Stock Producto Terminado (.xlsx)", type="xlsx", key="fspt")
    with col2:
        f_ppt = st.file_uploader("2. Producción de Hoy (.xlsx)", type="xlsx", key="fppt")

    if st.button("🚀 Sincronizar Producto Terminado"):
        if not f_spt:
            st.error("Sube al menos el archivo de stock.")
            st.stop()

        df_spt = leer_excel(f_spt, "Stock PT")
        if df_spt is None:
            st.stop()
        df_spt = normalizar_columnas(df_spt)
        faltan = columnas_faltantes(df_spt, ['Referencia', 'Descripcion', 'Cantidad'], "Stock PT")
        if faltan:
            st.stop()
        df_spt['Referencia'] = df_spt['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
        df_spt['Cantidad']   = pd.to_numeric(df_spt['Cantidad'], errors='coerce').fillna(0)

        # Guardar histórico en SQLite
        init_db()
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("DELETE FROM stock_pt_historico WHERE fecha = ?", (fecha_hoy,))
        for _, row in df_spt.iterrows():
            cur.execute("INSERT INTO stock_pt_historico (fecha, referencia, descripcion, cantidad) VALUES (?,?,?,?)",
                       (fecha_hoy, row['Referencia'], row.get('Descripcion',''), row['Cantidad']))
        con.commit()
        con.close()

        st.session_state.df_stock_pt = df_spt
        df_a_firebase(df_spt, 'producto_terminado', 'df_stock_pt')

        if f_ppt:
            df_ppt = leer_excel(f_ppt, "Producción PT")
            if df_ppt is not None:
                df_ppt = normalizar_columnas(df_ppt)
                faltan2 = columnas_faltantes(df_ppt, ['Referencia', 'Cantidad'], "Producción PT")
                if not faltan2:
                    df_ppt['Referencia'] = df_ppt['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                    df_ppt['Cantidad']   = pd.to_numeric(df_ppt['Cantidad'], errors='coerce').fillna(0)
                    st.session_state.df_produccion_pt = df_ppt
                    df_a_firebase(df_ppt, 'producto_terminado', 'df_produccion_pt')

        st.success("✅ Producto Terminado sincronizado.")
        st.rerun()

    if st.session_state.df_stock_pt is None:
        st.info("Sube el archivo de stock para empezar.")
        st.stop()

    # ── Dashboard ─────────────────────────────
    df_spt = st.session_state.df_stock_pt.copy()
    df_spt['Referencia'] = df_spt['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)

    # Producción de hoy
    if st.session_state.df_produccion_pt is not None:
        df_ppt = st.session_state.df_produccion_pt.copy()
        df_ppt['Referencia'] = df_ppt['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
        df_spt = df_spt.merge(df_ppt[['Referencia', 'Cantidad']].rename(columns={'Cantidad': 'Produccion'}),
                              on='Referencia', how='left')
        df_spt['Produccion'] = df_spt['Produccion'].fillna(0)
    else:
        df_spt['Produccion'] = 0

    # Ventas medias diarias desde df_ventas
    if st.session_state.df_ventas is not None:
        ventas = st.session_state.df_ventas.copy()
        ventas['Referencia'] = ventas['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
        ventas['Unidades']   = pd.to_numeric(ventas['Unidades'], errors='coerce').fillna(0)
        vmd = (ventas.groupby('Referencia')['Unidades'].sum() / DIAS_MES).reset_index()
        vmd.columns = ['Referencia', 'VMD']
        df_spt = df_spt.merge(vmd, on='Referencia', how='left')
        df_spt['VMD'] = df_spt['VMD'].fillna(0)
    else:
        df_spt['VMD'] = 0
        st.warning("⚠️ Sin datos de ventas. Carga el módulo de Etiquetas con el archivo de ventas.")

    df_spt['Stock_disponible'] = df_spt['Cantidad'] + df_spt['Produccion']
    df_spt['Dias_cobertura']   = df_spt.apply(
        lambda r: round(r['Stock_disponible'] / r['VMD'], 1) if r['VMD'] > 0 else 999, axis=1
    )

    def alerta_pt(row):
        if row['VMD'] == 0:
            return "⚪ Sin ventas", "#3d3d3d"
        if row['Dias_cobertura'] < 1:
            return "🔴 PELIGRO", "#721c24"
        if row['Dias_cobertura'] < 2:
            return "🟡 AVISO", "#856404"
        return "🟢 OK", "#155724"

    df_spt[['Estado', 'Color']] = df_spt.apply(lambda r: pd.Series(alerta_pt(r)), axis=1)

    # Métricas
    en_peligro = (df_spt['Estado'] == "🔴 PELIGRO").sum()
    en_aviso   = (df_spt['Estado'] == "🟡 AVISO").sum()
    mp1, mp2, mp3, mp4 = st.columns(4)
    mp1.metric("Total referencias", len(df_spt))
    mp2.metric("🔴 En peligro", en_peligro)
    mp3.metric("🟡 Aviso", en_aviso)
    mp4.metric("🟢 OK", (df_spt['Estado'] == "🟢 OK").sum())

    # Filtros
    filtro_pt = st.selectbox("Filtrar:", ["Todos", "🔴 Solo peligro", "🟡 Solo aviso", "🟢 Solo OK"], key="fpt")
    buscar_pt = st.text_input("Buscar referencia:", key="bpt")
    vista_pt = df_spt.copy()
    if filtro_pt == "🔴 Solo peligro":
        vista_pt = vista_pt[vista_pt['Estado'] == "🔴 PELIGRO"]
    elif filtro_pt == "🟡 Solo aviso":
        vista_pt = vista_pt[vista_pt['Estado'] == "🟡 AVISO"]
    elif filtro_pt == "🟢 Solo OK":
        vista_pt = vista_pt[vista_pt['Estado'] == "🟢 OK"]
    if buscar_pt:
        vista_pt = vista_pt[vista_pt['Referencia'].str.contains(buscar_pt, case=False, na=False)]

    cols_pt = ['Referencia', 'Descripcion', 'VMD', 'Cantidad', 'Produccion', 'Stock_disponible', 'Dias_cobertura', 'Estado']

    def colorear_pt(row):
        return [f'background-color: {vista_pt.loc[row.name, "Color"]}; color: white'] * len(row)

    st.dataframe(vista_pt[cols_pt].style.apply(colorear_pt, axis=1), use_container_width=True, height=500)

    import io
    out_pt = io.BytesIO()
    with pd.ExcelWriter(out_pt, engine='openpyxl') as w:
        vista_pt[cols_pt].to_excel(w, index=False, sheet_name='ProductoTerminado')
    out_pt.seek(0)
    st.download_button("📥 Exportar", out_pt, "producto_terminado.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════
# MÓDULO 11: PLANIFICACIÓN PRODUCCIÓN
# ══════════════════════════════════════════════
elif menu == "🏭 Planificación Producción":
    st.header("🏭 Planificación de la Producción")

    DIAS_MES = 22

    f_pp = st.file_uploader("Plan de Producción recibido (.xlsx)", type="xlsx", key="fpp")
    if f_pp and st.button("📥 Cargar Plan"):
        df_pp = leer_excel(f_pp, "Plan Producción")
        if df_pp is not None:
            df_pp = normalizar_columnas(df_pp)
            faltan = columnas_faltantes(df_pp, ['Referencia', 'Cantidad'], "Plan Producción")
            if not faltan:
                df_pp['Referencia'] = df_pp['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                df_pp['Cantidad']   = pd.to_numeric(df_pp['Cantidad'], errors='coerce').fillna(0)
                st.session_state.df_plan_produccion = df_pp
                df_a_firebase(df_pp, 'producto_terminado', 'df_plan_produccion')
                st.success("✅ Plan de producción cargado.")
                st.rerun()

    if st.session_state.df_stock_pt is None:
        st.warning("⚠️ Primero sincroniza el módulo de Producto Terminado.")
        st.stop()
    if st.session_state.df_ventas is None:
        st.warning("⚠️ Primero carga el archivo de ventas en el módulo de Etiquetas.")
        st.stop()

    df_spt = st.session_state.df_stock_pt.copy()
    df_spt['Referencia'] = df_spt['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
    df_spt['Cantidad']   = pd.to_numeric(df_spt['Cantidad'], errors='coerce').fillna(0)

    ventas = st.session_state.df_ventas.copy()
    ventas['Referencia'] = ventas['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
    ventas['Unidades']   = pd.to_numeric(ventas['Unidades'], errors='coerce').fillna(0)
    vmd = (ventas.groupby('Referencia')['Unidades'].sum() / DIAS_MES).reset_index()
    vmd.columns = ['Referencia', 'VMD']

    # Sugerencia: fabricar lo necesario para tener 1 día de stock
    df_sug = df_spt.merge(vmd, on='Referencia', how='left')
    df_sug['VMD'] = df_sug['VMD'].fillna(0)
    df_sug['Sugerencia'] = df_sug.apply(
        lambda r: max(0, math.ceil(r['VMD'] - r['Cantidad'])) if r['VMD'] > 0 else 0, axis=1
    )
    df_sug = df_sug[df_sug['VMD'] > 0].copy()

    # ── Turno de noche ────────────────────────
    st.subheader("🌙 Sugerencia Turno de Noche")
    st.caption("Referencias ordenadas por prioridad: mayor ratio venta/stock medio")

    # Stock medio histórico desde SQLite
    try:
        con = sqlite3.connect(DB_PATH)
        hist = pd.read_sql("SELECT referencia, AVG(cantidad) as stock_medio FROM stock_pt_historico GROUP BY referencia", con)
        con.close()
        hist['referencia'] = hist['referencia'].str.upper().str.strip()
        hist.columns = ['Referencia', 'Stock_medio']
        df_noche = df_sug.merge(hist, on='Referencia', how='left')
        df_noche['Stock_medio'] = df_noche['Stock_medio'].fillna(df_noche['Cantidad'])
    except Exception:
        df_noche = df_sug.copy()
        df_noche['Stock_medio'] = df_noche['Cantidad']

    df_noche['Prioridad'] = df_noche.apply(
        lambda r: round(r['VMD'] / max(r['Stock_medio'], 1), 4), axis=1
    )
    df_noche = df_noche[df_noche['Sugerencia'] > 0].sort_values('Prioridad', ascending=False)

    if df_noche.empty:
        st.success("✅ No hay referencias urgentes para fabricar esta noche.")
    else:
        st.dataframe(
            df_noche[['Referencia', 'Descripcion', 'VMD', 'Cantidad', 'Stock_medio', 'Sugerencia', 'Prioridad']]
            .reset_index(drop=True),
            use_container_width=True, height=400
        )

    st.divider()

    # ── Comparativa con plan recibido ─────────
    st.subheader("📋 Sugerencia vs Plan Recibido")
    if st.session_state.df_plan_produccion is not None:
        df_pp = st.session_state.df_plan_produccion.copy()
        df_pp['Referencia'] = df_pp['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
        df_pp = df_pp.rename(columns={'Cantidad': 'Plan_recibido'})

        comp = df_sug[['Referencia', 'Descripcion', 'VMD', 'Cantidad', 'Sugerencia']].merge(
            df_pp[['Referencia', 'Plan_recibido']], on='Referencia', how='outer'
        ).fillna(0)
        comp['Diferencia'] = comp['Sugerencia'] - comp['Plan_recibido']
        comp['Estado'] = comp['Diferencia'].apply(
            lambda d: f"🔴 Plan corto: {abs(int(d))} ud" if d > 0
            else (f"🟡 Plan largo: {abs(int(d))} ud" if d < 0 else "🟢 Coincide")
        )
        comp['Color'] = comp['Diferencia'].apply(
            lambda d: "#721c24" if d > 0 else ("#856404" if d < 0 else "#155724")
        )

        me1, me2, me3 = st.columns(3)
        me1.metric("🔴 Plan corto", (comp['Diferencia'] > 0).sum())
        me2.metric("🟡 Plan largo", (comp['Diferencia'] < 0).sum())
        me3.metric("🟢 Coincide", (comp['Diferencia'] == 0).sum())

        cols_comp = ['Referencia', 'Descripcion', 'VMD', 'Cantidad', 'Sugerencia', 'Plan_recibido', 'Diferencia', 'Estado']

        def colorear_comp(row):
            return [f'background-color: {comp.loc[row.name, "Color"]}; color: white'] * len(row)

        st.dataframe(comp[cols_comp].style.apply(colorear_comp, axis=1), use_container_width=True, height=400)

        import io
        out_comp = io.BytesIO()
        with pd.ExcelWriter(out_comp, engine='openpyxl') as w:
            comp[cols_comp].to_excel(w, index=False, sheet_name='Comparativa')
        out_comp.seek(0)
        st.download_button("📥 Exportar comparativa", out_comp, "plan_produccion.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Sube el plan de producción recibido para comparar con la sugerencia.")
        # Mostrar solo sugerencia
        st.dataframe(
            df_sug[['Referencia', 'Descripcion', 'VMD', 'Cantidad', 'Sugerencia']]
            .sort_values('Sugerencia', ascending=False)
            .reset_index(drop=True),
            use_container_width=True
        )

# ══════════════════════════════════════════════
# MÓDULO 12: LOGÍSTICA AI
# ══════════════════════════════════════════════
elif menu == "🧠 Logística AI":
    st.header("🧠 Logística AI")
    st.caption("Agente inteligente con memoria semántica — llama-3.3-70b-versatile")

    # ── Inicializar ChromaDB ──────────────────
    @st.cache_resource
    def get_chroma():
        import chromadb
        client = chromadb.PersistentClient(path="./memoria_logistica")
        col    = client.get_or_create_collection("logistica")
        return col

    coleccion_ai = get_chroma()

    def excel_a_chunks(df, nombre):
        chunks = []
        cols   = ", ".join(df.columns.tolist())
        chunks.append(f"Archivo: {nombre} | Columnas: {cols} | Total filas: {len(df)}")
        for i, row in df.head(500).iterrows():
            fila = f"[{nombre}] fila {i+1}: " + " | ".join(
                f"{col}={val}" for col, val in row.items() if pd.notna(val)
            )
            chunks.append(fila)
        return chunks

    def indexar_df(df, nombre):
        import time
        try:
            existentes = coleccion_ai.get(where={"archivo": nombre})["ids"]
            if existentes:
                coleccion_ai.delete(ids=existentes)
        except:
            pass
        chunks    = excel_a_chunks(df, nombre)
        ids       = [f"{nombre}_{i}_{int(time.time())}" for i in range(len(chunks))]
        metadatas = [{"archivo": nombre} for _ in chunks]
        coleccion_ai.add(documents=chunks, ids=ids, metadatas=metadatas)
        return len(chunks)

    def buscar_contexto_ai(pregunta, n=20):
        try:
            total = coleccion_ai.count()
            if total == 0:
                return ""
            resultados = coleccion_ai.query(query_texts=[pregunta], n_results=min(n, total))
            return "\n".join(resultados["documents"][0])
        except:
            return ""

    def parsear_medidas(texto):
        """Extrae LxAxH de strings como '600X400X162' o '18x25 36'"""
        import re
        partes = re.findall(r'[\d]+', str(texto))
        if len(partes) >= 3:
            return int(partes[0]), int(partes[1]), int(partes[2])
        return None, None, None

    # ── Panel lateral: archivos indexados ────
    col_main, col_side = st.columns([3, 1])

    with col_side:
        st.subheader("📁 Archivos indexados")

        # Indexar datos de la app automáticamente
        if st.button("🔄 Indexar datos de la app", use_container_width=True):
            with st.spinner("Indexando..."):
                n_total = 0
                if st.session_state.df_final is not None:
                    n_total += indexar_df(st.session_state.df_final, "maestro_bandejas")
                if st.session_state.df_consumos is not None:
                    n_total += indexar_df(st.session_state.df_consumos.head(1000), "consumos_bandejas")
                if st.session_state.df_materiales is not None:
                    n_total += indexar_df(st.session_state.df_materiales, "materiales_asociados")
                if st.session_state.df_etiquetas_final is not None:
                    n_total += indexar_df(st.session_state.df_etiquetas_final, "maestro_etiquetas")
                if st.session_state.df_ventas is not None:
                    n_total += indexar_df(st.session_state.df_ventas, "ventas_mensuales")
                if st.session_state.df_stock_pt is not None:
                    n_total += indexar_df(st.session_state.df_stock_pt, "stock_producto_terminado")
                if st.session_state.df_pedidos is not None and not st.session_state.df_pedidos.empty:
                    n_total += indexar_df(st.session_state.df_pedidos, "pedidos")
                st.session_state.logistica_archivos['app_datos'] = f"{n_total} chunks"
                st.success(f"✅ {n_total} chunks indexados")
                st.rerun()

        # Subir Excel adicional
        st.divider()
        f_extra = st.file_uploader("➕ Subir Excel adicional", type=["xlsx","csv"], key="fextra")
        if f_extra and st.button("📥 Indexar archivo", use_container_width=True):
            with st.spinner("Indexando..."):
                if f_extra.name.endswith('.csv'):
                    df_extra = pd.read_csv(f_extra)
                else:
                    df_extra = pd.read_excel(f_extra)
                df_extra = normalizar_columnas(df_extra)
                nombre_extra = f_extra.name.rsplit('.', 1)[0]
                n = indexar_df(df_extra, nombre_extra)
                st.session_state.logistica_archivos[nombre_extra] = f"{n} chunks | {len(df_extra)} filas"
                # Preguntar si es el archivo de paletización
                if st.checkbox(f"¿Es este el archivo de envases/paletización?", key=f"is_pal_{nombre_extra}"):
                    st.session_state.df_paletizacion = df_extra
                    df_a_firebase(df_extra, 'logistica', 'df_paletizacion')
                    st.info("📦 Guardado como archivo de paletización en Firebase")
                st.success(f"✅ {nombre_extra} indexado")
                st.rerun()

        # Mostrar archivos disponibles
        st.divider()
        try:
            total_docs = coleccion_ai.count()
            st.metric("Total chunks", total_docs)
        except:
            st.metric("Total chunks", 0)

        for nombre, info in st.session_state.logistica_archivos.items():
            st.caption(f"📄 {nombre}: {info}")

        if st.button("🗑️ Limpiar memoria", use_container_width=True):
            try:
                ids = coleccion_ai.get()["ids"]
                if ids:
                    coleccion_ai.delete(ids=ids)
            except:
                pass
            st.session_state.logistica_historial = []
            st.session_state.logistica_archivos  = {}
            st.success("Memoria limpiada")
            st.rerun()

    # ── Chat principal ────────────────────────
    with col_main:
        # Preguntas rápidas
        st.subheader("💡 Preguntas rápidas")
        preguntas = [
            "¿Qué referencias están en peligro hoy?",
            "¿Qué fabricamos esta noche?",
            "¿Qué etiquetas necesito pedir?",
            "Dame un resumen del estado del inventario",
            "¿Qué productos son obsoletos?",
        ]
        cols_p = st.columns(len(preguntas))
        for i, (col_p, preg) in enumerate(zip(cols_p, preguntas)):
            if col_p.button(preg, key=f"qai_{i}", use_container_width=True):
                st.session_state.logistica_historial.append({"role": "user", "content": preg})
                st.rerun()

        st.divider()

        # Historial del chat
        chat_container = st.container()
        with chat_container:
            for msg in st.session_state.logistica_historial:
                with st.chat_message(msg["role"]):
                    st.write(msg["content"])

        # Procesar si el último mensaje es del usuario
        if st.session_state.logistica_historial and st.session_state.logistica_historial[-1]["role"] == "user":
            pregunta_actual = st.session_state.logistica_historial[-1]["content"]
            with st.chat_message("assistant"):
                with st.spinner("Analizando..."):
                    try:
                        from groq import Groq as GroqClient
                        contexto = buscar_contexto_ai(pregunta_actual)
                        archivos_lista = ", ".join(st.session_state.logistica_archivos.keys()) or "ninguno aún"

                        # Enriquecer contexto con búsqueda exacta de referencias
                        import re as _re
                        refs_encontradas = _re.findall(r'[Cc]\d{4,6}|\d{5,6}', pregunta_actual)
                        contexto_extra = []

                        for ref in refs_encontradas:
                            ref_up = ref.upper().strip()
                            # Buscar en maestro bandejas
                            if st.session_state.df_final is not None:
                                df_b = st.session_state.df_final.copy()
                                df_b['Referencia'] = df_b['Referencia'].astype(str).str.upper().str.strip()
                                fila_b = df_b[df_b['Referencia'] == ref_up]
                                if not fila_b.empty:
                                    r = fila_b.iloc[0]
                                    import re as re2
                                    nums = re2.findall(r'\d+', str(r.get('Descripcion','')))
                                    medidas = ""
                                    if len(nums) >= 3:
                                        medidas = f"medidas: {int(nums[0])*10}x{int(nums[1])*10}x{nums[2]}mm"
                                    contexto_extra.append(f"BANDEJA {ref_up}: desc={r.get('Descripcion','')} | {medidas} | unidades_palet={r.get('Unidades_palet','')} | stock_interno={r.get('Stock_interno','')}")

                            # Buscar en maestro etiquetas
                            if st.session_state.df_etiquetas_final is not None:
                                df_e = st.session_state.df_etiquetas_final.copy()
                                df_e['Referencia'] = df_e['Referencia'].astype(str).str.upper().str.strip()
                                fila_e = df_e[df_e['Referencia'] == ref_up]
                                if not fila_e.empty:
                                    r = fila_e.iloc[0]
                                    contexto_extra.append(f"ETIQUETA {ref_up}: desc={r.get('Descripcion','')} | consumo_mes={r.get('Consumo_mes','')} | stock={r.get('Stock_interno','')}")

                            # Buscar en producto terminado
                            if st.session_state.df_stock_pt is not None:
                                df_pt = st.session_state.df_stock_pt.copy()
                                df_pt['Referencia'] = df_pt['Referencia'].astype(str).str.upper().str.strip().str.zfill(6)
                                ref_zf = ref_up.zfill(6)
                                fila_pt = df_pt[df_pt['Referencia'] == ref_zf]
                                if not fila_pt.empty:
                                    r = fila_pt.iloc[0]
                                    contexto_extra.append(f"PRODUCTO TERMINADO {ref_zf}: desc={r.get('Descripcion','')} | stock={r.get('Cantidad','')}")

                            # Buscar materiales asociados
                            if st.session_state.df_materiales is not None:
                                mat = st.session_state.df_materiales.copy()
                                mat['Referencia'] = mat['Referencia'].astype(str).str.upper().str.strip().str.zfill(6)
                                mat['Codigo'] = mat['Codigo'].astype(str).str.upper().str.strip()
                                ref_zf = ref_up.zfill(6) if ref_up.isdigit() else ref_up
                                # Como producto
                                comp = mat[mat['Referencia'] == ref_zf]
                                if not comp.empty:
                                    materiales = comp['Codigo'].tolist()
                                    contexto_extra.append(f"PRODUCTO {ref_zf} USA MATERIALES: {', '.join(materiales)}")
                                # Como material
                                prods = mat[mat['Codigo'] == ref_up]
                                if not prods.empty:
                                    productos = prods['Referencia'].tolist()
                                    contexto_extra.append(f"MATERIAL {ref_up} USADO EN PRODUCTOS: {', '.join(productos)}")

                        contexto_enriquecido = '\n'.join(contexto_extra) + '\n\n' + contexto if contexto_extra else contexto

                        # Busqueda por descripcion en DataFrames
                        palabras = [p for p in pregunta_actual.lower().split() if len(p) > 3
                                   and p not in ['cual','como','donde','tiene','cuanto','cuantos','cuantas',
                                                 'dame','dime','para','esta','este','tengo','quiero','busca',
                                                 'referencias','referencia','que','hay','del','los','las']]
                        if palabras:
                            desc_extra = []
                            for palabra in palabras[:3]:
                                if st.session_state.df_final is not None:
                                    df_bd = st.session_state.df_final.copy()
                                    mask = df_bd['Descripcion'].astype(str).str.lower().str.contains(palabra, na=False)
                                    for _, r in df_bd[mask].head(5).iterrows():
                                        desc_extra.append(f"BANDEJA ref={r['Referencia']} desc={r.get('Descripcion','')} uds_palet={r.get('Unidades_palet','')} stock_int={r.get('Stock_interno','')}")
                                if st.session_state.df_etiquetas_final is not None:
                                    df_ed = st.session_state.df_etiquetas_final.copy()
                                    mask = df_ed['Descripcion'].astype(str).str.lower().str.contains(palabra, na=False)
                                    for _, r in df_ed[mask].head(5).iterrows():
                                        desc_extra.append(f"ETIQUETA ref={r['Referencia']} desc={r.get('Descripcion','')} consumo_mes={r.get('Consumo_mes','')} stock={r.get('Stock_interno','')}")
                                if st.session_state.df_ventas is not None and 'Descripcion' in st.session_state.df_ventas.columns:
                                    df_vd = st.session_state.df_ventas.copy()
                                    df_vd['Referencia'] = df_vd['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                                    mask = df_vd['Descripcion'].astype(str).str.lower().str.contains(palabra, na=False)
                                    for _, r in df_vd[mask].drop_duplicates('Referencia').head(5).iterrows():
                                        desc_extra.append(f"VENTA ref={r['Referencia']} desc={r.get('Descripcion','')} uds={r.get('Unidades','')}")
                                        # Cruzar con materiales para encontrar bandeja
                                        if st.session_state.df_materiales is not None and st.session_state.df_final is not None:
                                            mat_vd = st.session_state.df_materiales.copy()
                                            mat_vd['Referencia'] = mat_vd['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                                            mat_vd['Codigo'] = mat_vd['Codigo'].astype(str).str.strip().str.upper()
                                            refs_bv = set(st.session_state.df_final['Referencia'].astype(str).str.strip().str.upper())
                                            df_bv = st.session_state.df_final.copy()
                                            df_bv['Referencia'] = df_bv['Referencia'].astype(str).str.strip().str.upper()
                                            import re as re_vd
                                            bandejas_vd = mat_vd[(mat_vd['Referencia'] == r['Referencia']) & (mat_vd['Codigo'].isin(refs_bv))]
                                            for _, bv in bandejas_vd.iterrows():
                                                fila_bv = df_bv[df_bv['Referencia'] == bv['Codigo']]
                                                if not fila_bv.empty:
                                                    rb = fila_bv.iloc[0]
                                                    nums_bv = re_vd.findall(r'\d+', str(rb.get('Descripcion','')))
                                                    medidas_bv = f"{int(nums_bv[0])*10}x{int(nums_bv[1])*10}x{nums_bv[2]}mm" if len(nums_bv)>=3 else ""
                                                    desc_extra.append(f"BANDEJA del producto {r['Referencia']}: ref={bv['Codigo']} desc={rb.get('Descripcion','')} medidas={medidas_bv} uds_palet={rb.get('Unidades_palet','')}")
                            if desc_extra:
                                # Cruzar productos encontrados con materiales para encontrar bandeja
                                if st.session_state.df_materiales is not None and st.session_state.df_final is not None:
                                    mat_ag = st.session_state.df_materiales.copy()
                                    mat_ag['Referencia'] = mat_ag['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                                    mat_ag['Codigo'] = mat_ag['Codigo'].astype(str).str.strip().str.upper()
                                    refs_band_ag = set(st.session_state.df_final['Referencia'].astype(str).str.strip().str.upper())
                                    df_band_ag = st.session_state.df_final.copy()
                                    df_band_ag['Referencia'] = df_band_ag['Referencia'].astype(str).str.strip().str.upper()
                                    for linea in desc_extra[:]:
                                        if linea.startswith('VENTA') or linea.startswith('ETIQUETA'):
                                            import re as re_ag
                                            ref_match = re_ag.search(r'ref=(\S+)', linea)
                                            if ref_match:
                                                ref_prod = ref_match.group(1).zfill(6)
                                                bandejas_prod = mat_ag[(mat_ag['Referencia'] == ref_prod) & (mat_ag['Codigo'].isin(refs_band_ag))]
                                                for _, bp in bandejas_prod.iterrows():
                                                    fila_bp = df_band_ag[df_band_ag['Referencia'] == bp['Codigo']]
                                                    if not fila_bp.empty:
                                                        rb = fila_bp.iloc[0]
                                                        nums_bp = re_ag.findall(r'\d+', str(rb.get('Descripcion','')))
                                                        medidas_bp = f"{int(nums_bp[0])*10}x{int(nums_bp[1])*10}x{nums_bp[2]}mm" if len(nums_bp)>=3 else ""
                                                        desc_extra.append(f"BANDEJA asociada al producto {ref_prod}: ref={bp['Codigo']} desc={rb.get('Descripcion','')} medidas={medidas_bp} uds_palet={rb.get('Unidades_palet','')}")
                                contexto_enriquecido = '\n'.join(desc_extra) + '\n\n' + contexto_enriquecido

                        system_prompt = f"""Eres un agente de logística inteligente de Aldelis, especializado en análisis de inventario, aprovisionamiento y planificación de producción.

Archivos disponibles: {archivos_lista}

DATOS EXACTOS DE REFERENCIAS MENCIONADAS:
{chr(10).join(contexto_extra) if contexto_extra else "No se encontraron referencias exactas en los datos."}

DATOS RELEVANTES ADICIONALES (búsqueda semántica):
{contexto if contexto else "No hay datos indexados."}

INSTRUCCIONES:
- Responde siempre en español, de forma concisa y práctica
- Usa los datos reales para responder con precisión
- Para paletizaciones: palet = 1200x800mm, altura máx ~2000mm. Bandejas: LxA en cm→mm (*10), alto en mm. Envases en mm
- Para calcular si una bandeja cabe en un envase: usa SIEMPRE las dimensiones INTERIORES del envase, que están en la columna "dimensiones interiores (mm)" del archivo "Paletizacion"
- Las dimensiones de todos los envases (IFCO, EUROPOOL, etc.) están en el archivo "Paletizacion" — búscalas ahí siempre
- Las bandejas SIEMPRE se colocan una encima de otra y una al lado de la otra. El cálculo correcto es: (largo_interior/largo_bandeja) x (ancho_interior/ancho_bandeja) x (alto_interior/alto_bandeja), redondeando hacia abajo en cada dimensión
- Nunca uses cálculo de volumen para paletizaciones, siempre usa dimensiones individuales
- Las dimensiones exteriores del envase NO se usan para calcular cuántas bandejas caben, solo las interiores
- Para turno de noche: prioriza referencias con mayor ratio venta_diaria/stock_medio
- Si cruzas datos de varios archivos, explícalo brevemente
- Si no encuentras la info, dilo claramente"""

                        groq_client = GroqClient(api_key=GROQ_API_KEY)
                        mensajes = [{"role": "system", "content": system_prompt}]
                        mensajes += st.session_state.logistica_historial[-10:]

                        resp = groq_client.chat.completions.create(
                            model="llama-3.3-70b-versatile",
                            messages=mensajes,
                            max_tokens=1500,
                            temperature=0.3
                        )
                        respuesta = resp.choices[0].message.content
                        st.write(respuesta)
                        st.session_state.logistica_historial.append({"role": "assistant", "content": respuesta})
                    except Exception as e:
                        st.error(f"Error: {e}")

        # Input
        if prompt := st.chat_input("Pregunta sobre logística, inventario, paletizaciones..."):
            st.session_state.logistica_historial.append({"role": "user", "content": prompt})
            st.rerun()

        # Limpiar conversación
        if st.session_state.logistica_historial:
            if st.button("🗑️ Limpiar conversación"):
                st.session_state.logistica_historial = []
                st.rerun()

        # ── Calculadora de Paletización ───
        st.divider()
        st.subheader("📦 Calculadora de Paletización")

        col_pal1, col_pal2, col_pal3 = st.columns(3)
        with col_pal1:
            ref_producto = st.text_input("Referencia producto (opcional):", key="pal_prod").strip().upper().zfill(6)
        with col_pal2:
            ref_bandeja = st.text_input("Referencia bandeja (ej: C12170):", key="pal_ref").strip().upper()
        with col_pal3:
            envases_disponibles = []
            if st.session_state.df_paletizacion is not None:
                df_pal_tmp = normalizar_columnas(st.session_state.df_paletizacion.copy())
                envases_disponibles = df_pal_tmp.iloc[:,0].astype(str).str.strip().tolist()
            ref_envase = st.selectbox("Selecciona envase:", [""] + envases_disponibles, key="pal_env")
            ref_envase = ref_envase.strip().upper() if ref_envase else ""

        if st.button("🔢 Calcular paletización"):
            import re as re_pal

            # Si se da producto, buscar bandeja en materiales asociados
            if ref_producto and ref_producto != '000000' and not ref_bandeja:
                if st.session_state.df_materiales is not None and st.session_state.df_final is not None:
                    mat_cp = st.session_state.df_materiales.copy()
                    mat_cp['Referencia'] = mat_cp['Referencia'].astype(str).str.strip().str.upper().str.zfill(6)
                    mat_cp['Codigo']     = mat_cp['Codigo'].astype(str).str.strip().str.upper()
                    refs_band = set(st.session_state.df_final['Referencia'].astype(str).str.strip().str.upper())
                    prod_mats = mat_cp[mat_cp['Referencia'] == ref_producto]
                    bandeja_encontrada = prod_mats[prod_mats['Codigo'].isin(refs_band)]
                    if not bandeja_encontrada.empty:
                        ref_bandeja = bandeja_encontrada.iloc[0]['Codigo']
                        st.info(f"🔍 Producto {ref_producto} usa bandeja: **{ref_bandeja}**")
                    else:
                        st.warning(f"No se encontró bandeja asociada al producto {ref_producto} en el maestro.")

            # Buscar medidas bandeja
            desc_band = None
            b_l = b_a = b_h = None
            if ref_bandeja and st.session_state.df_final is not None:
                df_b = st.session_state.df_final.copy()
                df_b['Referencia'] = df_b['Referencia'].astype(str).str.strip().str.upper()
                fila = df_b[df_b['Referencia'] == ref_bandeja]
                if not fila.empty:
                    desc_band = fila.iloc[0].get('Descripcion', '')
                    nums = re_pal.findall(r'\d+', str(desc_band))
                    if len(nums) >= 3:
                        b_l = int(nums[0]) * 10
                        b_a = int(nums[1]) * 10
                        b_h = int(nums[2])
                        st.session_state['pal_b_l'] = b_l
                        st.session_state['pal_b_a'] = b_a
                        st.session_state['pal_b_h'] = b_h
                        st.session_state['pal_ref_band'] = ref_bandeja

            # Buscar medidas envase directamente en df_paletizacion
            e_l = e_a = e_h = None
            uds_caja = None
            if ref_envase and st.session_state.df_paletizacion is not None:
                df_pal_bus = st.session_state.df_paletizacion.copy()
                df_pal_bus = normalizar_columnas(df_pal_bus)
                # Buscar columna de dimensiones interiores
                col_int = None
                for col in df_pal_bus.columns:
                    if 'inter' in col.lower():
                        col_int = col
                        break
                col_ext = None
                for col in df_pal_bus.columns:
                    if 'exter' in col.lower() or (col != col_int and ('dim' in col.lower() or 'mm' in col.lower())):
                        col_ext = col
                        break
                # Buscar por nombre de envase (búsqueda flexible)
                df_pal_bus['Envase_upper'] = df_pal_bus.iloc[:,0].astype(str).str.upper().str.strip()
                ref_env_upper = ref_envase.upper().strip()
                # Buscar coincidencia exacta primero, luego parcial
                fila_env = df_pal_bus[df_pal_bus['Envase_upper'] == ref_env_upper]
                if fila_env.empty:
                    fila_env = df_pal_bus[df_pal_bus['Envase_upper'].str.contains(ref_env_upper, na=False)]
                if not fila_env.empty and col_int:
                    dim_str = str(fila_env.iloc[0][col_int])
                    nums = re_pal.findall(r'\d+', dim_str)
                    if len(nums) >= 3:
                        e_l, e_a, e_h = int(nums[0]), int(nums[1]), int(nums[2])
                    # Numero de cajas
                    for col in df_pal_bus.columns:
                        if 'caja' in col.lower():
                            try:
                                uds_caja = int(fila_env.iloc[0][col])
                            except:
                                pass
                            break
                if e_l is None:
                    # Fallback a ChromaDB
                    ctx_env = buscar_contexto_ai(ref_envase, n=10)
                    st.warning(f"No se encontró '{ref_envase}' en el archivo de paletización. Buscando en ChromaDB...")

            if b_l and b_a and b_h:
                st.success(f"🏷️ Bandeja **{ref_bandeja}**: {b_l}×{b_a}×{b_h} mm")

                # ── CÁLCULO CON ENVASE ────────────────────
                if e_l and e_a and e_h:
                    st.success(f"📦 Envase **{ref_envase}** (interior): {e_l}×{e_a}×{e_h} mm")
                    HOLGURA_GAS = 5   # mm extra por capa por efecto del gas
                    MARGEN_CAJA = 4   # mm libres en la parte superior de la caja
                    b_h_real = b_h + HOLGURA_GAS
                    altura_util = e_h - MARGEN_CAJA

                    # Orientación normal: largo bandeja vs largo caja
                    fi_l_n = e_l // b_l
                    fi_a_n = e_a // b_a
                    fi_h_n = altura_util // b_h_real
                    uds_normal = fi_l_n * fi_a_n * fi_h_n

                    # Orientación girada: ancho bandeja vs largo caja
                    fi_l_g = e_l // b_a
                    fi_a_g = e_a // b_l
                    fi_h_g = altura_util // b_h_real
                    uds_girada = fi_l_g * fi_a_g * fi_h_g

                    # Elegir la mejor
                    if uds_girada > uds_normal:
                        fi_l, fi_a, fi_h = fi_l_g, fi_a_g, fi_h_g
                        uds_calculadas = uds_girada
                        orientacion = "↺ Girada (ancho×largo)"
                    else:
                        fi_l, fi_a, fi_h = fi_l_n, fi_a_n, fi_h_n
                        uds_calculadas = uds_normal
                        orientacion = "→ Normal (largo×ancho)"

                    # Mostrar ambas opciones
                    st.info(f"🔄 Orientación seleccionada: **{orientacion}** ({uds_calculadas} uds) | Alternativa: {uds_girada if orientacion.startswith('→') else uds_normal} uds")
                    st.subheader("📊 Resultado en caja")
                    cc1, cc2, cc3, cc4 = st.columns(4)
                    cc1.metric("Por fila (largo)", fi_l)
                    cc2.metric("Por fila (ancho)", fi_a)
                    cc3.metric("Capas en caja", fi_h)
                    cc4.metric("Total por caja", uds_calculadas)
                else:
                    st.warning(f"No se encontraron dimensiones para '{ref_envase}'. Asegúrate de que está indexado.")
                    uds_calculadas = uds_caja

                # ── CÁLCULO EN PALET ──────────────────────
                palet_l, palet_a, palet_h = 1200, 800, 2000
                st.subheader("🏗️ Resultado en palet (1200×800×2000mm)")

                if e_l and e_a and e_h:
                    # Cajas en palet
                    cajas_l = palet_l // e_l
                    cajas_a = palet_a // e_a
                    cajas_h = palet_h // e_h
                    cajas_palet = cajas_l * cajas_a * cajas_h
                    total_ud_palet = cajas_palet * uds_calculadas
                    cp1, cp2, cp3, cp4, cp5 = st.columns(5)
                    cp1.metric("Cajas/capa", cajas_l * cajas_a)
                    cp2.metric("Capas de cajas", cajas_h)
                    cp3.metric("Cajas/palet", cajas_palet)
                    cp4.metric("Uds/palet", total_ud_palet)
                    cp5.metric("Altura total", f"{cajas_h * e_h} mm")
                else:
                    # Sin envase, bandejas directo en palet
                    bp_l = palet_l // b_l
                    bp_a = palet_a // b_a
                    bp_h = palet_h // b_h
                    total_palet = bp_l * bp_a * bp_h
                    cp1, cp2, cp3, cp4 = st.columns(4)
                    cp1.metric("Por capa", bp_l * bp_a)
                    cp2.metric("Capas", bp_h)
                    cp3.metric("Total/palet", total_palet)
                    cp4.metric("Altura total", f"{bp_h * b_h} mm")
                    cajas_l = cajas_a = 1
                    cajas_h = bp_h
                    e_l = e_a = e_h = b_l
                    uds_calculadas = 1

            # SUGERENCIA IA DE ENVASE OPTIMO
            st.divider()
            # Recuperar medidas de session_state si el botón anterior no está activo
            if not (b_l and b_a and b_h):
                b_l = st.session_state.get('pal_b_l')
                b_a = st.session_state.get('pal_b_a')
                b_h = st.session_state.get('pal_b_h')
                ref_bandeja = st.session_state.get('pal_ref_band', ref_bandeja)
            if st.button("💡 Sugerir envase óptimo para esta bandeja") and b_l and b_a and b_h:
                if st.session_state.df_paletizacion is None:
                    st.warning("Sube el archivo de paletizaciones en este módulo primero.")
                else:
                    df_pal = st.session_state.df_paletizacion.copy()
                    df_pal = normalizar_columnas(df_pal)
                    import re as re_sug
                    resultados_envases = []
                    for _, row in df_pal.iterrows():
                        envase_nombre = str(row.iloc[0])
                        dim_int_col = None
                        for col in df_pal.columns:
                            if 'inter' in col.lower():
                                dim_int_col = col
                                break
                        if dim_int_col is None:
                            continue
                        dim_str = str(row.get(dim_int_col, ''))
                        nums = re_sug.findall(r'\d+', dim_str)
                        if len(nums) < 3:
                            continue
                        el, ea, eh = int(nums[0]), int(nums[1]), int(nums[2])
                        HOLGURA_GAS = 5   # mm extra por capa por efecto del gas
                        MARGEN_CAJA = 4   # mm libres en la parte superior de la caja
                        h_util = eh - MARGEN_CAJA
                        bh_r   = b_h + HOLGURA_GAS
                        # Probar ambas orientaciones
                        uds_n = (el // b_l) * (ea // b_a) * (h_util // bh_r)
                        uds_g = (el // b_a) * (ea // b_l) * (h_util // bh_r)
                        if uds_g > uds_n:
                            fi_l, fi_a, fi_h = el // b_a, ea // b_l, h_util // bh_r
                            uds = uds_g
                        else:
                            fi_l, fi_a, fi_h = el // b_l, ea // b_a, h_util // bh_r
                            uds = uds_n
                        if uds > 0:
                            resultados_envases.append({
                                'Envase': envase_nombre,
                                'Dim interior (mm)': f"{el}x{ea}x{eh}",
                                'Filas L': fi_l,
                                'Filas A': fi_a,
                                'Capas': fi_h,
                                'Uds/caja': uds,
                                'Aprovechamiento %': round(uds * b_l * b_a * b_h / (el * ea * eh) * 100, 1)
                            })
                    if resultados_envases:
                        df_res = pd.DataFrame(resultados_envases).sort_values('Uds/caja', ascending=False)
                        st.subheader(f"Envases compatibles con bandeja {ref_bandeja} ({b_l}x{b_a}x{b_h}mm)")
                        st.dataframe(df_res.reset_index(drop=True), use_container_width=True)
                        mejor = df_res.iloc[0]
                        tabla_str = df_res.to_string(index=False)
                        dim_band = f"{b_l}x{b_a}x{b_h}mm"
                        prompt_sug = f"Analiza estos envases para la bandeja {ref_bandeja} ({dim_band}) y da una recomendacion al comercial:\n\n{tabla_str}\n\nEl envase con mas unidades por caja es {mejor['Envase']} con {mejor['Uds/caja']} unidades. Explica brevemente las ventajas e inconvenientes de los 3 mejores y cual recomendarias y por que, en lenguaje sencillo para un comercial."
                        try:
                            from groq import Groq as GroqSug
                            groq_sug = GroqSug(api_key=GROQ_API_KEY)
                            resp_sug = groq_sug.chat.completions.create(
                                model="llama-3.3-70b-versatile",
                                messages=[
                                    {"role": "system", "content": "Eres un experto en logistica y packaging de Aldelis. Responde en espanol, de forma clara y practica para un comercial."},
                                    {"role": "user", "content": prompt_sug}
                                ],
                                max_tokens=800,
                                temperature=0.3
                            )
                            st.info("🤖 " + resp_sug.choices[0].message.content)
                        except Exception as e:
                            st.warning(f"No se pudo obtener sugerencia IA: {e}")
                    else:
                        st.warning("No se encontraron envases compatibles con esta bandeja.")

            elif ref_bandeja:
                st.warning(f"No se encontró la referencia {ref_bandeja} en el maestro de bandejas.")
            else:
                st.info("Introduce una referencia de bandeja o producto para calcular.")
